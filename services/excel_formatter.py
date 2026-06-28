"""
Excel Formatter Service for Project Aura.
Handles professional formatting for Excel workbooks.
"""

import logging
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelFormatter:
    """
    Service for applying professional formatting to Excel worksheets.
    """

    # Color scheme
    COLORS = {
        'header': '366092',      # Dark blue
        'header_text': 'FFFFFF',  # White
        'subheader': 'D9E1F2',    # Light blue
        'accent': '4472C4',       # Blue
        'success': 'C6E0B4',      # Light green
        'warning': 'FFE699',      # Light yellow
        'danger': 'F4B084',       # Light red
        'neutral': 'E2EFDA'       # Very light green
    }

    # Borders
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    @staticmethod
    def format_header_row(worksheet, row_num, columns):
        """
        Format a header row with professional styling.
        
        Args:
            worksheet: Excel worksheet
            row_num: Row number to format
            columns: List of column numbers (A=1, B=2, etc.)
        """
        for col in columns:
            cell = worksheet.cell(row=row_num, column=col)
            cell.fill = PatternFill(start_color=ExcelFormatter.COLORS['header'], 
                                   end_color=ExcelFormatter.COLORS['header'], 
                                   fill_type='solid')
            cell.font = Font(bold=True, color=ExcelFormatter.COLORS['header_text'], size=12)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = ExcelFormatter.THIN_BORDER

    @staticmethod
    def format_subheader_row(worksheet, row_num, columns):
        """
        Format a subheader row.
        
        Args:
            worksheet: Excel worksheet
            row_num: Row number to format
            columns: List of column numbers
        """
        for col in columns:
            cell = worksheet.cell(row=row_num, column=col)
            cell.fill = PatternFill(start_color=ExcelFormatter.COLORS['subheader'], 
                                   end_color=ExcelFormatter.COLORS['subheader'], 
                                   fill_type='solid')
            cell.font = Font(bold=True, size=11)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = ExcelFormatter.THIN_BORDER

    @staticmethod
    def format_data_row(worksheet, row_num, columns, fill_color=None):
        """
        Format a data row.
        
        Args:
            worksheet: Excel worksheet
            row_num: Row number
            columns: List of column numbers
            fill_color: Optional fill color
        """
        for col in columns:
            cell = worksheet.cell(row=row_num, column=col)
            if fill_color:
                cell.fill = PatternFill(start_color=fill_color, 
                                       end_color=fill_color, 
                                       fill_type='solid')
            cell.border = ExcelFormatter.THIN_BORDER
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    @staticmethod
    def set_column_widths(worksheet, column_widths):
        """
        Set column widths.
        
        Args:
            worksheet: Excel worksheet
            column_widths: Dict of {column_letter: width}
        """
        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width

    @staticmethod
    def freeze_panes(worksheet, freeze_row=1, freeze_col='A'):
        """
        Freeze panes for easy scrolling.
        
        Args:
            worksheet: Excel worksheet
            freeze_row: Row to freeze at
            freeze_col: Column to freeze at
        """
        freeze_cell = f'{freeze_col}{freeze_row + 1}'
        worksheet.freeze_panes = freeze_cell

    @staticmethod
    def add_filter(worksheet, start_cell, end_cell):
        """
        Add autofilter to a range.
        
        Args:
            worksheet: Excel worksheet
            start_cell: Starting cell (e.g., 'A1')
            end_cell: Ending cell (e.g., 'D10')
        """
        worksheet.auto_filter.ref = f'{start_cell}:{end_cell}'

    @staticmethod
    def format_title(worksheet, row, col, title):
        """
        Format a title cell.
        
        Args:
            worksheet: Excel worksheet
            row: Row number
            col: Column number
            title: Title text
        """
        cell = worksheet.cell(row=row, column=col)
        cell.value = title
        cell.font = Font(bold=True, size=14, color='FFFFFF')
        cell.fill = PatternFill(start_color=ExcelFormatter.COLORS['header'], 
                               end_color=ExcelFormatter.COLORS['header'], 
                               fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    @staticmethod
    def format_label(worksheet, row, col, label):
        """
        Format a label cell.
        
        Args:
            worksheet: Excel worksheet
            row: Row number
            col: Column number
            label: Label text
        """
        cell = worksheet.cell(row=row, column=col)
        cell.value = label
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal='left', vertical='center')

    @staticmethod
    def format_value(worksheet, row, col, value, bold=False, color=None):
        """
        Format a value cell.
        
        Args:
            worksheet: Excel worksheet
            row: Row number
            col: Column number
            value: Cell value
            bold: Whether to bold
            color: Optional fill color
        """
        cell = worksheet.cell(row=row, column=col)
        cell.value = value
        cell.font = Font(bold=bold, size=10)
        if color:
            cell.fill = PatternFill(start_color=color, 
                                   end_color=color, 
                                   fill_type='solid')
        cell.alignment = Alignment(horizontal='left', vertical='center')

    @staticmethod
    def format_percentage_column(worksheet, start_row, end_row, col):
        """
        Format a column as percentages.
        
        Args:
            worksheet: Excel worksheet
            start_row: Starting row
            end_row: Ending row
            col: Column number
        """
        for row in range(start_row, end_row + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.number_format = '0%'

    @staticmethod
    def format_currency_column(worksheet, start_row, end_row, col):
        """
        Format a column as currency.
        
        Args:
            worksheet: Excel worksheet
            start_row: Starting row
            end_row: Ending row
            col: Column number
        """
        for row in range(start_row, end_row + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.number_format = '$#,##0.00'

    @staticmethod
    def format_date_column(worksheet, start_row, end_row, col):
        """
        Format a column as dates.
        
        Args:
            worksheet: Excel worksheet
            start_row: Starting row
            end_row: Ending row
            col: Column number
        """
        for row in range(start_row, end_row + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.number_format = 'mm/dd/yyyy'

    @staticmethod
    def set_sheet_color(worksheet, color):
        """
        Set the sheet tab color.
        
        Args:
            worksheet: Excel worksheet
            color: Color code (without #)
        """
        worksheet.sheet_properties.tabColor = color

    @staticmethod
    def merge_cells(worksheet, start_cell, end_cell):
        """
        Merge cells.
        
        Args:
            worksheet: Excel worksheet
            start_cell: Starting cell (e.g., 'A1')
            end_cell: Ending cell (e.g., 'C1')
        """
        worksheet.merge_cells(f'{start_cell}:{end_cell}')

    @staticmethod
    def add_shape_style(worksheet, row, col, cell_type='status'):
        """
        Add conditional formatting or styling.
        
        Args:
            worksheet: Excel worksheet
            row: Row number
            col: Column number
            cell_type: Type of cell ('status', 'priority', etc.)
        """
        cell = worksheet.cell(row=row, column=col)
        
        if cell_type == 'status':
            value = str(cell.value).lower()
            if 'completed' in value or 'done' in value:
                cell.fill = PatternFill(start_color=ExcelFormatter.COLORS['success'], 
                                       end_color=ExcelFormatter.COLORS['success'], 
                                       fill_type='solid')
            elif 'in progress' in value or 'active' in value:
                cell.fill = PatternFill(start_color=ExcelFormatter.COLORS['warning'], 
                                       end_color=ExcelFormatter.COLORS['warning'], 
                                       fill_type='solid')
            elif 'pending' in value or 'blocked' in value:
                cell.fill = PatternFill(start_color=ExcelFormatter.COLORS['danger'], 
                                       end_color=ExcelFormatter.COLORS['danger'], 
                                       fill_type='solid')
        
        cell.border = ExcelFormatter.THIN_BORDER

    @staticmethod
    def format_severity(worksheet, row, col, severity):
        """
        Format a severity cell with color coding.
        
        Args:
            worksheet: Excel worksheet
            row: Row number
            col: Column number
            severity: Severity level ('High', 'Medium', 'Low')
        """
        cell = worksheet.cell(row=row, column=col)
        cell.value = severity
        cell.font = Font(bold=True, size=10)
        
        if severity == 'High':
            cell.fill = PatternFill(start_color=ExcelFormatter.COLORS['danger'], 
                                   end_color=ExcelFormatter.COLORS['danger'], 
                                   fill_type='solid')
        elif severity == 'Medium':
            cell.fill = PatternFill(start_color=ExcelFormatter.COLORS['warning'], 
                                   end_color=ExcelFormatter.COLORS['warning'], 
                                   fill_type='solid')
        else:  # Low
            cell.fill = PatternFill(start_color=ExcelFormatter.COLORS['success'], 
                                   end_color=ExcelFormatter.COLORS['success'], 
                                   fill_type='solid')
        
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = ExcelFormatter.THIN_BORDER
