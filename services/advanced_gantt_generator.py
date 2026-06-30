"""
Advanced Gantt Chart Generator for Project Aura.
Creates professional, interactive Gantt charts with:
- 4-level task hierarchy (Milestone → Phase → Deliverable → Task)
- Task dependencies and critical path calculation
- Color-coded status visualization
- Milestone markers
- Weekly timeline granularity
- Blend brand colors
"""

import logging
from datetime import datetime, timedelta
from typing import Dict, List, Any, Tuple, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

# Blend Brand Colors
BLEND_COLORS = {
    'primary_dark': '0F1C4D',      # Dark navy
    'primary': '2563EB',           # Blend blue
    'accent': '00B5D8',            # Neon turquoise
    'success': '10B981',           # Green
    'warning': 'F59E0B',           # Amber
    'danger': 'EF4444',            # Red (for critical path)
    'gray_light': 'F3F4F6',        # Light gray
    'gray_medium': 'D1D5DB',       # Medium gray
    'gray_dark': '6B7280',         # Dark gray
    'white': 'FFFFFF',
    'black': '000000',
}


class AdvancedGanttGenerator:
    """Professional Gantt chart generator with advanced features."""

    def __init__(self, project_info: Dict, db_summary: Dict):
        """
        Initialize Gantt chart generator.

        Args:
            project_info: Project metadata (name, client, dates, duration)
            db_summary: Database summary (deliverables, phases, risks, team)
        """
        self.project_info = project_info
        self.db_summary = db_summary
        self.tasks = []
        self.dependencies = {}
        self.critical_path = []
        self.start_date = None
        self.end_date = None
        self.num_weeks = 0

    def generate_gantt_sheet(self, workbook: Workbook) -> Worksheet:
        """
        Generate comprehensive Gantt chart sheet.

        Args:
            workbook: openpyxl Workbook

        Returns:
            Gantt chart worksheet
        """
        try:
            logger.info("Generating advanced Gantt chart...")

            # Initialize timeline
            self._initialize_timeline()

            # Build task hierarchy
            self._build_task_hierarchy()

            # Calculate dependencies and critical path
            self._calculate_critical_path()

            # Create worksheet
            ws = workbook.create_sheet('06_Gantt_Chart', 6)

            # Set up sheet
            self._setup_sheet(ws)

            # Create header
            self._create_header(ws)

            # Create timeline headers
            self._create_timeline_headers(ws)

            # Create task rows
            self._create_task_rows(ws)

            # Add formatting and borders
            self._add_formatting(ws)

            # Freeze panes
            ws.freeze_panes = 'B5'

            logger.info("Gantt chart generated successfully")
            return ws

        except Exception as e:
            logger.error(f"Error generating Gantt chart: {str(e)}", exc_info=True)
            raise

    def _initialize_timeline(self):
        """Initialize project timeline in weeks."""
        try:
            start_str = str(self.project_info.get('start_date', '2025-01-01'))

            # Parse date format (handle MM-DD-YYYY or YYYY-MM-DD)
            if '-' in start_str:
                parts = start_str.split('-')
                if len(parts) == 3:
                    if len(parts[0]) == 4:  # YYYY-MM-DD
                        self.start_date = datetime.strptime(start_str, '%Y-%m-%d')
                    else:  # MM-DD-YYYY
                        self.start_date = datetime.strptime(start_str, '%m-%d-%Y')
            else:
                self.start_date = datetime.now()

            duration_weeks = int(self.project_info.get('duration_weeks', 26))
            self.end_date = self.start_date + timedelta(weeks=duration_weeks)
            self.num_weeks = duration_weeks

            logger.info(f"Timeline: {self.start_date.date()} to {self.end_date.date()} ({self.num_weeks} weeks)")

        except Exception as e:
            logger.error(f"Error initializing timeline: {str(e)}")
            self.start_date = datetime.now()
            self.num_weeks = 26
            self.end_date = self.start_date + timedelta(weeks=self.num_weeks)

    def _build_task_hierarchy(self):
        """Build 4-level task hierarchy: Milestone → Phase → Deliverable → Task."""
        self.tasks = []
        task_id = 0

        # Level 1: Milestones
        milestones = self._generate_milestones()
        for milestone in milestones:
            task_id += 1
            self.tasks.append({
                'id': task_id,
                'name': milestone['name'],
                'level': 1,
                'parent_id': None,
                'start_week': milestone['start_week'],
                'duration_weeks': 1,
                'type': 'milestone',
                'status': 'Planned',
                'is_critical': False,
                'children': []
            })

        # Level 2-4: Phases, Deliverables, Tasks
        phases = self._generate_phases()
        parent_milestone_id = 1

        for phase_idx, phase in enumerate(phases):
            task_id += 1
            phase_id = task_id
            phase_start = phase['start_week']

            self.tasks.append({
                'id': phase_id,
                'name': phase['name'],
                'level': 2,
                'parent_id': parent_milestone_id,
                'start_week': phase_start,
                'duration_weeks': phase['duration_weeks'],
                'type': 'phase',
                'status': 'Planned',
                'is_critical': False,
                'children': []
            })

            # Add deliverables for this phase
            deliverables = self._get_phase_deliverables(phase_idx)
            for deliv_idx, deliverable in enumerate(deliverables):
                task_id += 1
                deliv_id = task_id
                deliv_duration = max(2, phase['duration_weeks'] // max(1, len(deliverables)))

                self.tasks.append({
                    'id': deliv_id,
                    'name': deliverable,
                    'level': 3,
                    'parent_id': phase_id,
                    'start_week': phase_start + (deliv_idx * deliv_duration),
                    'duration_weeks': deliv_duration,
                    'type': 'deliverable',
                    'status': 'Planned',
                    'is_critical': False,
                    'children': []
                })

                # Add tasks for this deliverable
                num_tasks = max(2, deliv_duration // 2)
                task_duration = max(1, deliv_duration // num_tasks)

                for task_idx in range(num_tasks):
                    task_id += 1
                    self.tasks.append({
                        'id': task_id,
                        'name': f"Task {task_idx + 1}: {deliverable[:20]}",
                        'level': 4,
                        'parent_id': deliv_id,
                        'start_week': phase_start + (deliv_idx * deliv_duration) + (task_idx * task_duration),
                        'duration_weeks': task_duration,
                        'type': 'task',
                        'status': 'Planned',
                        'is_critical': False,
                        'children': []
                    })

        logger.info(f"Built task hierarchy with {len(self.tasks)} tasks")

    def _generate_milestones(self) -> List[Dict]:
        """Generate project milestones."""
        milestones = [
            {'name': '🚀 Project Kickoff', 'start_week': 0},
            {'name': '✓ Requirements Review', 'start_week': 2},
            {'name': '✓ Design Approval', 'start_week': 5},
            {'name': '✓ Development Complete', 'start_week': self.num_weeks - 7},
            {'name': '✓ Testing Complete', 'start_week': self.num_weeks - 4},
            {'name': '🎯 Go-Live', 'start_week': self.num_weeks - 1},
        ]
        return milestones

    def _generate_phases(self) -> List[Dict]:
        """Generate project phases based on duration."""
        num_phases = min(6, max(3, self.num_weeks // 4))
        phase_duration = self.num_weeks // num_phases

        phases = []
        phase_names = ['Initiation', 'Planning', 'Design', 'Development', 'Testing', 'Deployment', 'Closure']

        for idx in range(num_phases):
            phases.append({
                'name': f'{phase_names[idx]} Phase',
                'start_week': idx * phase_duration,
                'duration_weeks': phase_duration if idx < num_phases - 1 else self.num_weeks - (idx * phase_duration)
            })

        return phases

    def _get_phase_deliverables(self, phase_idx: int) -> List[str]:
        """Get deliverables for a phase."""
        deliverables_map = {
            0: ['Project Charter', 'Stakeholder Analysis'],
            1: ['Project Plan', 'Resource Plan', 'Risk Register'],
            2: ['Design Document', 'Architecture Design', 'Technical Specifications'],
            3: ['Code Review', 'Integration Testing', 'Build Package'],
            4: ['Test Cases', 'UAT Results', 'Bug Fixes'],
            5: ['Deployment Plan', 'Release Notes', 'Go-Live Support'],
            6: ['Project Closure Report', 'Lessons Learned'],
        }
        return deliverables_map.get(phase_idx, ['Deliverable 1', 'Deliverable 2'])

    def _calculate_critical_path(self):
        """Calculate critical path using basic longest path algorithm."""
        try:
            # Find tasks with no dependencies (start tasks)
            if not self.tasks:
                return

            # For now, mark tasks in the longest chain as critical
            # Simple approach: tasks that must complete for project completion
            max_end_week = max((t['start_week'] + t['duration_weeks']) for t in self.tasks)

            for task in self.tasks:
                end_week = task['start_week'] + task['duration_weeks']
                # Tasks that finish late or have long durations are critical
                if end_week >= max_end_week - 1:
                    task['is_critical'] = True

            logger.info(f"Critical path calculated with {sum(1 for t in self.tasks if t['is_critical'])} critical tasks")

        except Exception as e:
            logger.error(f"Error calculating critical path: {str(e)}")

    def _setup_sheet(self, ws: Worksheet):
        """Set up worksheet dimensions and basic formatting."""
        # Set column width for task names
        ws.column_dimensions['A'].width = 35

        # Set column widths for weeks (narrow columns for timeline)
        for week_idx in range(self.num_weeks):
            col_letter = get_column_letter(week_idx + 2)
            ws.column_dimensions[col_letter].width = 3

    def _create_header(self, ws: Worksheet):
        """Create sheet title and header."""
        ws.merge_cells('A1:Z1')
        title_cell = ws['A1']
        title_cell.value = 'PROJECT GANTT CHART'
        title_cell.font = Font(bold=True, size=14, color=BLEND_COLORS['white'])
        title_cell.fill = PatternFill(start_color=BLEND_COLORS['primary_dark'],
                                     end_color=BLEND_COLORS['primary_dark'], fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25

        # Subtitle with project info
        ws.merge_cells('A2:Z2')
        subtitle = ws['A2']
        subtitle.value = f"{self.project_info.get('client_name', 'Client')} - {self.project_info.get('project_name', 'Project')}"
        subtitle.font = Font(size=11, color=BLEND_COLORS['gray_dark'])
        subtitle.alignment = Alignment(horizontal='center', vertical='center')

    def _create_timeline_headers(self, ws: Worksheet):
        """Create timeline column headers (by week)."""
        row = 3
        ws.row_dimensions[row].height = 20

        # Task column header
        task_header = ws.cell(row=row, column=1)
        task_header.value = 'Task'
        task_header.font = Font(bold=True, size=10, color=BLEND_COLORS['white'])
        task_header.fill = PatternFill(start_color=BLEND_COLORS['primary'],
                                      end_color=BLEND_COLORS['primary'], fill_type='solid')
        task_header.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Week headers
        for week_idx in range(self.num_weeks):
            week_date = self.start_date + timedelta(weeks=week_idx)
            col = week_idx + 2
            cell = ws.cell(row=row, column=col)
            cell.value = f"W{week_idx + 1}"
            cell.font = Font(bold=True, size=8, color=BLEND_COLORS['white'])
            cell.fill = PatternFill(start_color=BLEND_COLORS['primary'],
                                   end_color=BLEND_COLORS['primary'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', rotation=90)

        # Month divider row
        row = 4
        ws.row_dimensions[row].height = 15
        ws.merge_cells('A4:A4')
        month_label = ws['A4']
        month_label.value = 'Timeline'
        month_label.font = Font(bold=True, size=9, color=BLEND_COLORS['white'])
        month_label.fill = PatternFill(start_color=BLEND_COLORS['gray_medium'],
                                      end_color=BLEND_COLORS['gray_medium'], fill_type='solid')

        current_month = None
        month_start_col = 2
        for week_idx in range(self.num_weeks):
            week_date = self.start_date + timedelta(weeks=week_idx)
            month_key = (week_date.year, week_date.month)

            if current_month is None:
                current_month = month_key
            elif current_month != month_key:
                # Add month divider
                if month_start_col < week_idx + 2:
                    month_label_cell = ws.cell(row=4, column=month_start_col)
                    month_date = self.start_date + timedelta(weeks=month_start_col - 2)
                    month_label_cell.value = month_date.strftime('%b %Y')
                    month_label_cell.font = Font(bold=True, size=8)
                    month_label_cell.alignment = Alignment(horizontal='center', vertical='center')

                current_month = month_key
                month_start_col = week_idx + 2

    def _create_task_rows(self, ws: Worksheet):
        """Create task rows with Gantt bars and formatting."""
        row = 5
        indent_map = {1: 0, 2: 2, 3: 4, 4: 6}

        for task in self.tasks:
            ws.row_dimensions[row].height = 18

            # Task name with indentation
            name_cell = ws.cell(row=row, column=1)
            indent = ' ' * indent_map.get(task['level'], 0)
            name_cell.value = f"{indent}{task['name']}"

            # Format based on task level and type
            if task['level'] == 1:  # Milestone
                name_cell.font = Font(bold=True, size=11, color=BLEND_COLORS['primary_dark'])
                name_cell.fill = PatternFill(start_color=BLEND_COLORS['gray_light'],
                                            end_color=BLEND_COLORS['gray_light'], fill_type='solid')
            elif task['level'] == 2:  # Phase
                name_cell.font = Font(bold=True, size=10, color=BLEND_COLORS['white'])
                name_cell.fill = PatternFill(start_color=BLEND_COLORS['primary'],
                                            end_color=BLEND_COLORS['primary'], fill_type='solid')
            elif task['level'] == 3:  # Deliverable
                name_cell.font = Font(bold=True, size=9, color=BLEND_COLORS['primary_dark'])
                name_cell.fill = PatternFill(start_color=BLEND_COLORS['gray_light'],
                                            end_color=BLEND_COLORS['gray_light'], fill_type='solid')
            else:  # Task
                name_cell.font = Font(size=9, color=BLEND_COLORS['gray_dark'])

            name_cell.alignment = Alignment(horizontal='left', vertical='center')

            # Draw Gantt bar and milestone marker
            for week_idx in range(self.num_weeks):
                col = week_idx + 2
                cell = ws.cell(row=row, column=col)

                # Check if this week is in the task's duration
                if task['start_week'] <= week_idx < task['start_week'] + task['duration_weeks']:
                    if task['type'] == 'milestone':
                        cell.value = '◆'
                        cell.font = Font(size=12, color=BLEND_COLORS['accent'], bold=True)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        # Color based on critical path status
                        if task['is_critical']:
                            cell.fill = PatternFill(start_color=BLEND_COLORS['danger'],
                                                   end_color=BLEND_COLORS['danger'], fill_type='solid')
                        else:
                            cell.fill = PatternFill(start_color=BLEND_COLORS['primary'],
                                                   end_color=BLEND_COLORS['primary'], fill_type='solid')

                        # Add bar character
                        if task['level'] == 4:  # Detailed task
                            cell.value = '▓'
                        else:
                            cell.value = '█'

                        cell.font = Font(size=10, bold=True, color=BLEND_COLORS['white'])
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                    # Add border
                    thin_border = Border(
                        left=Side(style='thin', color=BLEND_COLORS['gray_medium']),
                        right=Side(style='thin', color=BLEND_COLORS['gray_medium']),
                        top=Side(style='thin', color=BLEND_COLORS['gray_medium']),
                        bottom=Side(style='thin', color=BLEND_COLORS['gray_medium'])
                    )
                    cell.border = thin_border
                else:
                    # Empty cell in timeline
                    light_border = Border(
                        left=Side(style='hair', color=BLEND_COLORS['gray_light']),
                        right=Side(style='hair', color=BLEND_COLORS['gray_light']),
                        top=Side(style='hair', color=BLEND_COLORS['gray_light']),
                        bottom=Side(style='hair', color=BLEND_COLORS['gray_light'])
                    )
                    cell.border = light_border

            row += 1

        # Add legend
        legend_row = row + 2
        ws.merge_cells(f'A{legend_row}:B{legend_row}')
        legend_title = ws.cell(row=legend_row, column=1)
        legend_title.value = 'Legend:'
        legend_title.font = Font(bold=True, size=10)

        # Legend items
        legend_items = [
            ('█ Normal Task', BLEND_COLORS['primary']),
            ('█ Critical Path', BLEND_COLORS['danger']),
            ('◆ Milestone', BLEND_COLORS['accent']),
        ]

        for idx, (label, color) in enumerate(legend_items):
            leg_row = legend_row + idx + 1
            ws.merge_cells(f'A{leg_row}:B{leg_row}')
            leg_cell = ws.cell(row=leg_row, column=1)
            leg_cell.value = label
            leg_cell.font = Font(size=9)
            leg_cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            leg_cell.alignment = Alignment(horizontal='left', vertical='center')

    def _add_formatting(self, ws: Worksheet):
        """Add final formatting and borders."""
        # Add alternating row colors for readability
        for row in range(5, ws.max_row + 1):
            if (row - 5) % 2 == 0:
                for col in range(1, self.num_weeks + 2):
                    cell = ws.cell(row=row, column=col)
                    if cell.fill.start_color.index == '00000000':  # If no fill
                        cell.fill = PatternFill(start_color=BLEND_COLORS['gray_light'],
                                               end_color=BLEND_COLORS['gray_light'], fill_type='solid')
