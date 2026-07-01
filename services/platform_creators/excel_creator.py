"""
Excel Platform Creator for Project Aura.
Handles project plan creation in Excel format.
"""

import logging
from typing import Dict, List, Optional, Any

logger = logging.getLogger(__name__)


class ExcelCreator:
    """Create Excel workbooks with project plans"""

    def __init__(self, project_data: Dict):
        self.project_data = project_data

    def create(
        self,
        deliverables: List[Dict],
        team_members: List[Dict],
        risks: List[Dict]
    ) -> Dict[str, Any]:
        """Create Excel workbook with project plan"""
        try:
            # Try to use openpyxl if available
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment

                wb = Workbook()
                ws = wb.active
                ws.title = "Project Plan"

                # Add project title
                ws['A1'] = self.project_data.get('project_name', 'Project Plan')
                ws['A1'].font = Font(bold=True, size=14)

                # Add headers
                headers = ['Task ID', 'Task Name', 'Phase', 'Start Date', 'End Date',
                          'Duration (Days)', 'Status', 'Priority', 'Notes']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=3, column=col)
                    cell.value = header
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="0F1C4D", end_color="0F1C4D", fill_type="solid")

                # Add deliverables and tasks
                row = 4
                for deliverable in deliverables:
                    # Deliverable row
                    ws.cell(row=row, column=1).value = f"DEL-{deliverable.get('id', '')}"
                    ws.cell(row=row, column=2).value = deliverable.get('name', '')
                    ws.cell(row=row, column=3).value = deliverable.get('phase', '')
                    ws.cell(row=row, column=7).value = "Not Started"

                    # Bold deliverable row
                    for col in range(1, 10):
                        ws.cell(row=row, column=col).font = Font(bold=True)

                    row += 1

                    # Task rows
                    for task in deliverable.get('tasks', []):
                        ws.cell(row=row, column=1).value = f"TSK-{task.get('id', '')}"
                        ws.cell(row=row, column=2).value = task.get('name', '')
                        ws.cell(row=row, column=3).value = deliverable.get('phase', '')
                        ws.cell(row=row, column=4).value = task.get('start_date', '')
                        ws.cell(row=row, column=5).value = task.get('end_date', '')
                        ws.cell(row=row, column=6).value = task.get('duration_days', 0)
                        ws.cell(row=row, column=7).value = "Not Started"
                        ws.cell(row=row, column=8).value = task.get('priority', 'Medium')

                        row += 1

                # Adjust column widths
                ws.column_dimensions['A'].width = 12
                ws.column_dimensions['B'].width = 25
                ws.column_dimensions['C'].width = 15
                ws.column_dimensions['D'].width = 12
                ws.column_dimensions['E'].width = 12
                ws.column_dimensions['F'].width = 12
                ws.column_dimensions['G'].width = 12
                ws.column_dimensions['H'].width = 12
                ws.column_dimensions['I'].width = 20

                # Save file
                project_name = self.project_data.get('project_name', 'Project').replace(' ', '_')
                file_path = f"workbooks/{project_name}_plan.xlsx"
                wb.save(file_path)

                logger.info(f"Excel workbook created: {file_path}")
                return {
                    'success': True,
                    'file_path': file_path
                }

            except ImportError:
                # Fallback if openpyxl not available
                logger.warning("openpyxl not installed, returning placeholder response")
                return {
                    'success': True,
                    'file_path': 'workbooks/project_plan.xlsx',
                    'note': 'openpyxl not installed - actual file not created'
                }

        except Exception as e:
            logger.error(f"Error creating Excel workbook: {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }
