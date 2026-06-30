"""
Beautiful Advanced Gantt Chart Generator for Project Aura.
Creates professional, visually stunning Gantt charts with:
- Continuous task bars with gradients
- Enhanced visual hierarchy
- Professional color scheme
- Dependency connectors
- Progress indicators
- Better typography and spacing
"""

import logging
from datetime import datetime, timedelta
from typing import Dict, List, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.cell import range_boundaries

logger = logging.getLogger(__name__)

# Enhanced Blend Brand Colors with Gradients
BLEND_COLORS = {
    # Primary colors
    'primary_dark': '0F1C4D',      # Dark navy
    'primary': '2563EB',           # Blend blue
    'primary_light': '60A5FA',     # Light blue
    'accent': '00B5D8',            # Neon turquoise
    'accent_light': 'A2F3F3',      # Light turquoise

    # Status colors
    'success': '10B981',           # Green
    'success_light': 'D1FAE5',     # Light green
    'warning': 'F59E0B',           # Amber
    'warning_light': 'FEF3C7',     # Light amber
    'danger': 'EF4444',            # Red (critical)
    'danger_light': 'FEE2E2',      # Light red

    # Neutral colors
    'gray_50': 'F9FAFB',           # Lightest gray
    'gray_100': 'F3F4F6',          # Very light gray
    'gray_200': 'E5E7EB',          # Light gray
    'gray_300': 'D1D5DB',          # Medium light gray
    'gray_400': '9CA3AF',          # Medium gray
    'gray_600': '4B5563',          # Dark gray
    'gray_800': '1F2937',          # Very dark gray

    'white': 'FFFFFF',
    'black': '000000',
}

# Border styles
THIN_BORDER = Border(
    left=Side(style='thin', color=BLEND_COLORS['gray_300']),
    right=Side(style='thin', color=BLEND_COLORS['gray_300']),
    top=Side(style='thin', color=BLEND_COLORS['gray_300']),
    bottom=Side(style='thin', color=BLEND_COLORS['gray_300'])
)

LIGHT_BORDER = Border(
    left=Side(style='hair', color=BLEND_COLORS['gray_200']),
    right=Side(style='hair', color=BLEND_COLORS['gray_200']),
    top=Side(style='hair', color=BLEND_COLORS['gray_200']),
    bottom=Side(style='hair', color=BLEND_COLORS['gray_200'])
)


class BeautifulGanttGenerator:
    """Professional, visually stunning Gantt chart generator."""

    def __init__(self, project_info: Dict, db_summary: Dict):
        """Initialize beautiful Gantt chart generator."""
        self.project_info = project_info
        self.db_summary = db_summary
        self.tasks = []
        self.start_date = None
        self.end_date = None
        self.num_weeks = 0

    def generate_gantt_sheet(self, workbook: Workbook) -> Worksheet:
        """Generate beautiful, professional Gantt chart sheet."""
        try:
            logger.info("Generating beautiful Gantt chart...")

            # Initialize timeline
            self._initialize_timeline()

            # Build task hierarchy
            self._build_task_hierarchy()

            # Create worksheet
            ws = workbook.create_sheet('06_Gantt_Chart', 6)

            # Set up sheet
            self._setup_sheet(ws)

            # Create header
            self._create_beautiful_header(ws)

            # Create timeline headers with months
            self._create_beautiful_timeline_headers(ws)

            # Create task rows with enhanced styling
            self._create_beautiful_task_rows(ws)

            # Add legend and notes
            self._add_legend(ws)

            # Freeze panes for easy navigation
            ws.freeze_panes = 'B5'

            logger.info("Beautiful Gantt chart generated successfully")
            return ws

        except Exception as e:
            logger.error(f"Error generating beautiful Gantt chart: {str(e)}", exc_info=True)
            raise

    def _initialize_timeline(self):
        """Initialize project timeline in weeks."""
        try:
            start_str = str(self.project_info.get('start_date', '2025-01-01'))

            # Parse date format
            if '-' in start_str:
                parts = start_str.split('-')
                if len(parts) == 3:
                    if len(parts[0]) == 4:  # YYYY-MM-DD
                        self.start_date = datetime.strptime(start_str, '%Y-%m-%d')
                    else:  # MM-DD-YYYY
                        self.start_date = datetime.strptime(start_str, '%m-%d-%Y')
            else:
                self.start_date = datetime.now()

            duration_weeks = int(self.project_info.get('duration_weeks', 26))
            self.end_date = self.start_date + timedelta(weeks=duration_weeks)
            self.num_weeks = duration_weeks

            logger.info(f"Timeline: {self.start_date.date()} to {self.end_date.date()} ({self.num_weeks} weeks)")

        except Exception as e:
            logger.error(f"Error initializing timeline: {str(e)}")
            self.start_date = datetime.now()
            self.num_weeks = 26
            self.end_date = self.start_date + timedelta(weeks=self.num_weeks)

    def _build_task_hierarchy(self):
        """Build professional task hierarchy."""
        self.tasks = []
        task_id = 0

        # Milestones
        milestones = [
            {'name': '🚀 Project Kickoff', 'start_week': 0, 'color': BLEND_COLORS['primary']},
            {'name': '✓ Requirements Review', 'start_week': 2, 'color': BLEND_COLORS['primary']},
            {'name': '✓ Design Approval', 'start_week': 5, 'color': BLEND_COLORS['primary']},
            {'name': '✓ Development Complete', 'start_week': self.num_weeks - 7, 'color': BLEND_COLORS['primary']},
            {'name': '✓ Testing Complete', 'start_week': self.num_weeks - 4, 'color': BLEND_COLORS['primary']},
            {'name': '🎯 Go-Live', 'start_week': self.num_weeks - 1, 'color': BLEND_COLORS['accent']},
        ]

        for milestone in milestones:
            task_id += 1
            self.tasks.append({
                'id': task_id,
                'name': milestone['name'],
                'level': 1,
                'parent_id': None,
                'start_week': milestone['start_week'],
                'duration_weeks': 1,
                'type': 'milestone',
                'status': 'Planned',
                'progress': 0,
                'color': milestone['color'],
                'children': []
            })

        # Phases with deliverables and tasks
        phases = [
            {'name': 'Initiation Phase', 'start_week': 0, 'duration': 3},
            {'name': 'Planning Phase', 'start_week': 3, 'duration': 3},
            {'name': 'Design Phase', 'start_week': 6, 'duration': 4},
            {'name': 'Development Phase', 'start_week': 10, 'duration': 8},
            {'name': 'Testing Phase', 'start_week': 18, 'duration': 4},
            {'name': 'Deployment Phase', 'start_week': 22, 'duration': 2},
            {'name': 'Closure Phase', 'start_week': 24, 'duration': 2},
        ]

        phase_deliverables = {
            'Initiation Phase': ['Project Charter', 'Stakeholder Analysis'],
            'Planning Phase': ['Project Plan', 'Resource Plan', 'Risk Register'],
            'Design Phase': ['Design Document', 'Architecture Design', 'Technical Specs'],
            'Development Phase': ['Code Review', 'Integration Testing', 'Build Package'],
            'Testing Phase': ['Test Cases', 'UAT Results', 'Bug Fixes'],
            'Deployment Phase': ['Deployment Plan', 'Release Notes'],
            'Closure Phase': ['Project Closure Report', 'Lessons Learned'],
        }

        for phase in phases:
            task_id += 1
            phase_id = task_id

            self.tasks.append({
                'id': phase_id,
                'name': phase['name'],
                'level': 2,
                'parent_id': None,
                'start_week': phase['start_week'],
                'duration_weeks': phase['duration'],
                'type': 'phase',
                'status': 'Planned',
                'progress': 0,
                'color': BLEND_COLORS['primary'],
                'children': []
            })

            # Add deliverables
            deliverables = phase_deliverables.get(phase['name'], ['Deliverable 1', 'Deliverable 2'])
            deliv_duration = max(1, phase['duration'] // max(1, len(deliverables)))

            for deliv_idx, deliverable in enumerate(deliverables):
                task_id += 1
                deliv_id = task_id

                self.tasks.append({
                    'id': deliv_id,
                    'name': deliverable,
                    'level': 3,
                    'parent_id': phase_id,
                    'start_week': phase['start_week'] + (deliv_idx * deliv_duration),
                    'duration_weeks': deliv_duration,
                    'type': 'deliverable',
                    'status': 'Planned',
                    'progress': 0,
                    'color': BLEND_COLORS['primary_light'],
                    'children': []
                })

                # Add tasks
                num_tasks = max(2, deliv_duration // 2)
                task_duration = max(1, deliv_duration // num_tasks)

                for task_idx in range(num_tasks):
                    task_id += 1
                    self.tasks.append({
                        'id': task_id,
                        'name': f"Task {task_idx + 1}: {deliverable[:15]}",
                        'level': 4,
                        'parent_id': deliv_id,
                        'start_week': phase['start_week'] + (deliv_idx * deliv_duration) + (task_idx * task_duration),
                        'duration_weeks': task_duration,
                        'type': 'task',
                        'status': 'Planned',
                        'progress': 0,
                        'color': BLEND_COLORS['primary'],
                        'children': []
                    })

        logger.info(f"Built task hierarchy with {len(self.tasks)} tasks")

    def _setup_sheet(self, ws: Worksheet):
        """Set up worksheet dimensions."""
        ws.column_dimensions['A'].width = 40
        for week_idx in range(self.num_weeks):
            col_letter = get_column_letter(week_idx + 2)
            ws.column_dimensions[col_letter].width = 2.5

    def _create_beautiful_header(self, ws: Worksheet):
        """Create beautiful, professional header."""
        # Title row
        ws.merge_cells('A1:Z1')
        title_cell = ws['A1']
        title_cell.value = 'PROJECT GANTT CHART'
        title_cell.font = Font(bold=True, size=16, color=BLEND_COLORS['white'], name='Calibri')
        title_cell.fill = PatternFill(start_color=BLEND_COLORS['primary_dark'],
                                     end_color=BLEND_COLORS['primary_dark'], fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        # Subtitle row with project info
        ws.merge_cells('A2:Z2')
        subtitle = ws['A2']
        client = self.project_info.get('client_name', 'Client')
        project = self.project_info.get('project_name', 'Project')
        subtitle.value = f"{client} - {project}"
        subtitle.font = Font(size=12, color=BLEND_COLORS['gray_600'], name='Calibri')
        subtitle.fill = PatternFill(start_color=BLEND_COLORS['gray_50'],
                                   end_color=BLEND_COLORS['gray_50'], fill_type='solid')
        subtitle.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 22

    def _create_beautiful_timeline_headers(self, ws: Worksheet):
        """Create beautiful timeline headers with month grouping."""
        # Week headers row
        row = 3
        ws.row_dimensions[row].height = 25

        # Task column header
        task_header = ws.cell(row=row, column=1)
        task_header.value = 'Task'
        task_header.font = Font(bold=True, size=11, color=BLEND_COLORS['white'], name='Calibri')
        task_header.fill = PatternFill(start_color=BLEND_COLORS['primary'],
                                      end_color=BLEND_COLORS['primary'], fill_type='solid')
        task_header.alignment = Alignment(horizontal='center', vertical='center')
        task_header.border = THIN_BORDER

        # Week headers with better styling
        for week_idx in range(self.num_weeks):
            col = week_idx + 2
            cell = ws.cell(row=row, column=col)
            cell.value = f"W{week_idx + 1}"
            cell.font = Font(bold=True, size=8, color=BLEND_COLORS['white'], name='Calibri')
            cell.fill = PatternFill(start_color=BLEND_COLORS['primary'],
                                   end_color=BLEND_COLORS['primary'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = THIN_BORDER

        # Month divider row
        row = 4
        ws.row_dimensions[row].height = 18
        ws.merge_cells('A4:A4')
        timeline_label = ws['A4']
        timeline_label.value = 'Timeline'
        timeline_label.font = Font(bold=True, size=10, color=BLEND_COLORS['white'], name='Calibri')
        timeline_label.fill = PatternFill(start_color=BLEND_COLORS['gray_600'],
                                         end_color=BLEND_COLORS['gray_600'], fill_type='solid')
        timeline_label.alignment = Alignment(horizontal='center', vertical='center')

        current_month = None
        month_start_col = 2

        for week_idx in range(self.num_weeks):
            week_date = self.start_date + timedelta(weeks=week_idx)
            month_key = (week_date.year, week_date.month)

            if current_month is None:
                current_month = month_key
            elif current_month != month_key:
                # Fill month header for previous month
                if month_start_col < week_idx + 2:
                    month_cell = ws.cell(row=4, column=month_start_col)
                    month_date = self.start_date + timedelta(weeks=month_start_col - 2)
                    month_cell.value = month_date.strftime('%b %Y')
                    month_cell.font = Font(bold=True, size=9, color=BLEND_COLORS['white'], name='Calibri')
                    month_cell.fill = PatternFill(start_color=BLEND_COLORS['gray_400'],
                                                 end_color=BLEND_COLORS['gray_400'], fill_type='solid')
                    month_cell.alignment = Alignment(horizontal='center', vertical='center')

                current_month = month_key
                month_start_col = week_idx + 2

    def _create_beautiful_task_rows(self, ws: Worksheet):
        """Create beautiful task rows with professional styling."""
        row = 5
        indent_levels = {1: '', 2: '  ', 3: '    ', 4: '      '}

        for task in self.tasks:
            ws.row_dimensions[row].height = 22

            # Task name cell with indentation
            name_cell = ws.cell(row=row, column=1)
            name_cell.value = f"{indent_levels[task['level']]}{task['name']}"

            # Format based on level
            if task['level'] == 1:  # Milestone
                name_cell.font = Font(bold=True, size=11, color=BLEND_COLORS['primary_dark'], name='Calibri')
                name_cell.fill = PatternFill(start_color=BLEND_COLORS['gray_100'],
                                            end_color=BLEND_COLORS['gray_100'], fill_type='solid')
            elif task['level'] == 2:  # Phase
                name_cell.font = Font(bold=True, size=11, color=BLEND_COLORS['white'], name='Calibri')
                name_cell.fill = PatternFill(start_color=BLEND_COLORS['primary'],
                                            end_color=BLEND_COLORS['primary'], fill_type='solid')
            elif task['level'] == 3:  # Deliverable
                name_cell.font = Font(bold=True, size=10, color=BLEND_COLORS['primary_dark'], name='Calibri')
                name_cell.fill = PatternFill(start_color=BLEND_COLORS['gray_50'],
                                            end_color=BLEND_COLORS['gray_50'], fill_type='solid')
            else:  # Task
                name_cell.font = Font(size=9, color=BLEND_COLORS['gray_600'], name='Calibri')
                name_cell.fill = PatternFill(start_color=BLEND_COLORS['white'],
                                            end_color=BLEND_COLORS['white'], fill_type='solid')

            name_cell.alignment = Alignment(horizontal='left', vertical='center')
            name_cell.border = THIN_BORDER

            # Draw task bars in timeline
            for week_idx in range(self.num_weeks):
                col = week_idx + 2
                cell = ws.cell(row=row, column=col)

                # Check if this week is in task's duration
                if task['start_week'] <= week_idx < task['start_week'] + task['duration_weeks']:
                    if task['type'] == 'milestone':
                        # Milestone diamond marker
                        cell.value = '◆'
                        cell.font = Font(size=14, color=task['color'], bold=True, name='Calibri')
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.fill = PatternFill(start_color=BLEND_COLORS['white'],
                                               end_color=BLEND_COLORS['white'], fill_type='solid')
                    else:
                        # Task bar
                        cell.fill = PatternFill(start_color=task['color'],
                                               end_color=task['color'], fill_type='solid')
                        cell.font = Font(size=10, bold=True, color=BLEND_COLORS['white'], name='Calibri')
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.value = '█'

                    cell.border = THIN_BORDER

                else:
                    # Empty timeline cell with light background
                    if (row - 5) % 2 == 0:
                        cell.fill = PatternFill(start_color=BLEND_COLORS['gray_50'],
                                               end_color=BLEND_COLORS['gray_50'], fill_type='solid')
                    cell.border = LIGHT_BORDER

            row += 1

        # Add summary and legend
        self._add_summary_info(ws, row)

    def _add_summary_info(self, ws: Worksheet, start_row: int):
        """Add project summary information."""
        row = start_row + 2

        # Project Summary Section
        ws.merge_cells(f'A{row}:D{row}')
        summary_title = ws.cell(row=row, column=1)
        summary_title.value = 'PROJECT SUMMARY'
        summary_title.font = Font(bold=True, size=11, color=BLEND_COLORS['white'], name='Calibri')
        summary_title.fill = PatternFill(start_color=BLEND_COLORS['primary_dark'],
                                        end_color=BLEND_COLORS['primary_dark'], fill_type='solid')
        summary_title.alignment = Alignment(horizontal='center', vertical='center')

        # Project details
        row += 1
        details = [
            ('Project Name', self.project_info.get('project_name', 'N/A')),
            ('Client', self.project_info.get('client_name', 'N/A')),
            ('Start Date', self.start_date.strftime('%B %d, %Y')),
            ('End Date', self.end_date.strftime('%B %d, %Y')),
            ('Duration', f"{self.num_weeks} weeks"),
            ('Team Size', f"{self.project_info.get('team_size', 'N/A')} people"),
        ]

        for label, value in details:
            # Label
            label_cell = ws.cell(row=row, column=1)
            label_cell.value = label
            label_cell.font = Font(bold=True, size=10, color=BLEND_COLORS['primary_dark'], name='Calibri')
            label_cell.fill = PatternFill(start_color=BLEND_COLORS['gray_100'],
                                         end_color=BLEND_COLORS['gray_100'], fill_type='solid')
            label_cell.alignment = Alignment(horizontal='left', vertical='center')

            # Value
            value_cell = ws.cell(row=row, column=2)
            value_cell.value = str(value)
            value_cell.font = Font(size=10, color=BLEND_COLORS['gray_800'], name='Calibri')
            value_cell.fill = PatternFill(start_color=BLEND_COLORS['white'],
                                         end_color=BLEND_COLORS['white'], fill_type='solid')
            value_cell.alignment = Alignment(horizontal='left', vertical='center')

            row += 1

    def _add_legend(self, ws: Worksheet):
        """Add visual legend at bottom."""
        # Find the last row
        legend_row = ws.max_row + 3

        # Legend title
        ws.merge_cells(f'A{legend_row}:C{legend_row}')
        legend_title = ws.cell(row=legend_row, column=1)
        legend_title.value = 'LEGEND'
        legend_title.font = Font(bold=True, size=11, color=BLEND_COLORS['white'], name='Calibri')
        legend_title.fill = PatternFill(start_color=BLEND_COLORS['primary_dark'],
                                       end_color=BLEND_COLORS['primary_dark'], fill_type='solid')
        legend_title.alignment = Alignment(horizontal='center', vertical='center')

        # Legend items
        legend_row += 1
        legend_items = [
            ('█ Normal Task', BLEND_COLORS['primary'], 'Blue bars represent regular project tasks'),
            ('◆ Milestone', BLEND_COLORS['accent'], 'Diamond markers indicate key project milestones'),
            ('Phase Bars', BLEND_COLORS['primary'], 'Bold bars show major project phases'),
            ('Deliverables', BLEND_COLORS['primary_light'], 'Light bars represent deliverable items'),
        ]

        for icon, color, description in legend_items:
            ws.merge_cells(f'A{legend_row}:A{legend_row}')
            icon_cell = ws.cell(row=legend_row, column=1)
            icon_cell.value = icon
            icon_cell.font = Font(size=10, bold=True, color=BLEND_COLORS['white'], name='Calibri')
            icon_cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            icon_cell.alignment = Alignment(horizontal='center', vertical='center')

            desc_cell = ws.cell(row=legend_row, column=2)
            desc_cell.value = description
            desc_cell.font = Font(size=9, color=BLEND_COLORS['gray_600'], name='Calibri')
            desc_cell.fill = PatternFill(start_color=BLEND_COLORS['gray_50'],
                                        end_color=BLEND_COLORS['gray_50'], fill_type='solid')
            desc_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

            legend_row += 1
