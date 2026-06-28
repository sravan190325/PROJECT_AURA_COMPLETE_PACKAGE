"""
Enhanced Workbook Generation Service for Project Aura.
Adds executive-grade project management sheets and professional formatting.
"""

import logging
from datetime import datetime, timedelta
from typing import Dict, Any, List
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, DoughnutChart, Reference
from openpyxl.utils import get_column_letter

from services.excel_formatter import ExcelFormatter
from services.project_plan_engine import ProjectPlanEngine

logger = logging.getLogger(__name__)


class WorkbookEnhancements:
    """Enhanced worksheet generation with executive-grade formatting."""

    # Extended color palette for enterprise look
    COLOR_PALETTE = {
        'primary_dark': '1F4E78',
        'primary': '366092',
        'primary_light': 'D9E1F2',
        'success': '70AD47',
        'warning': 'FFC000',
        'danger': 'C5504F',
        'info': '4472C4',
        'neutral': 'BFBFBF',
        'white': 'FFFFFF',
        'black': '000000'
    }

    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    @staticmethod
    def create_executive_dashboard(workbook, project_info: Dict, db_summary: Dict):
        """Create Executive Dashboard sheet with KPIs and health indicators."""
        ws = workbook.create_sheet('00_Executive_Dashboard', 0)

        # Title
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = f"{project_info.get('client_name', 'Client')} - Executive Project Dashboard"
        title_cell.font = Font(bold=True, size=16, color=WorkbookEnhancements.COLOR_PALETTE['white'])
        title_cell.fill = PatternFill(start_color=WorkbookEnhancements.COLOR_PALETTE['primary_dark'],
                                     end_color=WorkbookEnhancements.COLOR_PALETTE['primary_dark'],
                                     fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        # Generation timestamp
        ws.merge_cells('A2:H2')
        ts_cell = ws['A2']
        ts_cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ts_cell.font = Font(size=10, italic=True, color='666666')
        ts_cell.alignment = Alignment(horizontal='right')

        # KPI Cards section
        row = 4
        kpi_data = [
            ('Project Name', project_info.get('project_name', 'N/A')),
            ('Client Name', project_info.get('client_name', 'N/A')),
            ('Project Type', project_info.get('project_type', 'Unknown')),
            ('Start Date', project_info.get('start_date', 'TBD')),
            ('End Date', WorkbookEnhancements._calculate_end_date(project_info), ),
            ('Duration', f"{project_info.get('duration_weeks', 0)} weeks"),
            ('Team Size', project_info.get('team_size', 0)),
            ('Status', 'On Track'),
        ]

        for label, value in kpi_data:
            ws.merge_cells(f'A{row}:B{row}')
            label_cell = ws[f'A{row}']
            label_cell.value = label
            label_cell.font = Font(bold=True, size=11, color=WorkbookEnhancements.COLOR_PALETTE['white'])
            label_cell.fill = PatternFill(start_color=WorkbookEnhancements.COLOR_PALETTE['primary'],
                                         end_color=WorkbookEnhancements.COLOR_PALETTE['primary'],
                                         fill_type='solid')
            label_cell.alignment = Alignment(horizontal='left', vertical='center')

            ws.merge_cells(f'C{row}:D{row}')
            value_cell = ws[f'C{row}']
            value_cell.value = value
            value_cell.font = Font(size=11, bold=True)
            value_cell.fill = PatternFill(start_color=WorkbookEnhancements.COLOR_PALETTE['primary_light'],
                                         end_color=WorkbookEnhancements.COLOR_PALETTE['primary_light'],
                                         fill_type='solid')
            value_cell.alignment = Alignment(horizontal='left', vertical='center')

            row += 1

        # Health Indicator
        row += 1
        health = WorkbookEnhancements._calculate_project_health(project_info, db_summary)

        ws.merge_cells(f'A{row}:B{row}')
        health_label = ws[f'A{row}']
        health_label.value = 'Project Health'
        health_label.font = Font(bold=True, size=11, color=WorkbookEnhancements.COLOR_PALETTE['white'])
        health_label.fill = PatternFill(start_color=WorkbookEnhancements.COLOR_PALETTE['primary'],
                                       end_color=WorkbookEnhancements.COLOR_PALETTE['primary'],
                                       fill_type='solid')

        ws.merge_cells(f'C{row}:D{row}')
        health_cell = ws[f'C{row}']
        health_cell.value = health['status']
        health_cell.font = Font(size=11, bold=True, color=WorkbookEnhancements.COLOR_PALETTE['white'])
        health_cell.fill = PatternFill(start_color=health['color'],
                                      end_color=health['color'],
                                      fill_type='solid')
        health_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Summary metrics
        row += 3
        ws[f'A{row}'] = 'SUMMARY METRICS'
        ws[f'A{row}'].font = Font(bold=True, size=12)

        row += 1
        metrics = [
            ('Risks', len(db_summary.get('risks', []))),
            ('Team Members', len(db_summary.get('team_members', []))),
            ('Deliverables', len(db_summary.get('deliverables', []))),
        ]

        for metric, value in metrics:
            ws[f'A{row}'] = metric
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            ws[f'B{row}'].font = Font(bold=True, size=11)
            row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        for col in ['E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 15

        ExcelFormatter.freeze_panes(ws, 3)

    @staticmethod
    def create_project_roadmap(workbook, project_info: Dict):
        """Create Project Roadmap sheet with phase-level timeline."""
        ws = workbook.create_sheet('01_Project_Roadmap')

        # Title
        ExcelFormatter.format_title(ws, 1, 1, 'Project Roadmap')
        ws.merge_cells('A1:F1')

        # Headers
        row = 3
        headers = ['Phase', 'Start Date', 'End Date', 'Duration (Weeks)', 'Status', 'Key Deliverables']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            ExcelFormatter.format_header_row(ws, row, [col])

        # Generate phases
        row += 1
        plan_data = ProjectPlanEngine.generate_project_plan(project_info)
        start_date = datetime.strptime(project_info.get('start_date', '2025-01-01'), '%Y-%m-%d') if isinstance(project_info.get('start_date'), str) else project_info.get('start_date')

        for phase in plan_data.get('phases', []):
            ExcelFormatter.format_data_row(ws, row, list(range(1, 7)))
            ws.cell(row=row, column=1).value = phase.get('phase', '')
            ws.cell(row=row, column=2).value = start_date.strftime('%Y-%m-%d')
            end = start_date + timedelta(days=phase.get('duration_days', 7))
            ws.cell(row=row, column=3).value = end.strftime('%Y-%m-%d')
            weeks = phase.get('duration_days', 0) / 5
            ws.cell(row=row, column=4).value = round(weeks, 1)
            ws.cell(row=row, column=5).value = 'Planned'
            ws.cell(row=row, column=6).value = ''
            start_date = end + timedelta(days=1)
            row += 1

        ExcelFormatter.set_column_widths(ws, {'A': 20, 'B': 15, 'C': 15, 'D': 15, 'E': 15, 'F': 30})
        ExcelFormatter.freeze_panes(ws, 3)

    @staticmethod
    def create_enhanced_project_plan(workbook, project_info: Dict, db_summary: Dict):
        """Create Enhanced Project Plan with detailed task structure."""
        ws = workbook.create_sheet('02_Enhanced_Project_Plan', 2)

        ExcelFormatter.format_title(ws, 1, 1, 'Detailed Project Plan')
        ws.merge_cells('A1:M1')

        # Headers
        row = 3
        headers = ['Phase', 'Task', 'Deliverable', 'Owner', 'Start Date', 'End Date',
                   'Duration (Days)', 'Dependency', 'Status', 'Completion %', 'Priority', 'Risk Level', 'Notes']
        for col, header in enumerate(headers, 1):
            ExcelFormatter.format_header_row(ws, row, [col])
            ws.cell(row=row, column=col).value = header

        row += 1
        plan_data = ProjectPlanEngine.generate_project_plan(project_info)

        for phase in plan_data.get('phases', []):
            # Phase header row
            ws.merge_cells(f'A{row}:M{row}')
            phase_cell = ws[f'A{row}']
            phase_cell.value = phase.get('phase', '').upper()
            phase_cell.font = Font(bold=True, size=11, color=WorkbookEnhancements.COLOR_PALETTE['white'])
            phase_cell.fill = PatternFill(start_color=WorkbookEnhancements.COLOR_PALETTE['info'],
                                         end_color=WorkbookEnhancements.COLOR_PALETTE['info'],
                                         fill_type='solid')
            row += 1

            # Task rows
            for _ in range(2):  # 2 tasks per phase
                ExcelFormatter.format_data_row(ws, row, list(range(1, 14)))
                ws.cell(row=row, column=1).value = phase.get('phase', '')
                ws.cell(row=row, column=2).value = f"Task {_+1}"
                ws.cell(row=row, column=3).value = ''
                ws.cell(row=row, column=4).value = 'PM'
                ws.cell(row=row, column=5).value = phase.get('start_date', '')
                ws.cell(row=row, column=6).value = phase.get('end_date', '')
                ws.cell(row=row, column=7).value = phase.get('duration_days', 0)
                ws.cell(row=row, column=8).value = ''
                ws.cell(row=row, column=9).value = 'Planned'
                ws.cell(row=row, column=10).value = 0
                ExcelFormatter.format_percentage_column(ws, row, row, 10)
                ws.cell(row=row, column=11).value = 'Medium'
                ws.cell(row=row, column=12).value = 'Low'
                row += 1

        ExcelFormatter.set_column_widths(ws, {
            'A': 15, 'B': 15, 'C': 15, 'D': 12, 'E': 12, 'F': 12, 'G': 12, 'H': 12,
            'I': 10, 'J': 12, 'K': 10, 'L': 12, 'M': 15
        })
        ExcelFormatter.freeze_panes(ws, 4)

    @staticmethod
    def create_milestone_tracker(workbook, project_info: Dict, db_summary: Dict):
        """Create Milestone Tracker sheet."""
        ws = workbook.create_sheet('03_Milestone_Tracker')

        ExcelFormatter.format_title(ws, 1, 1, 'Milestone Tracker')
        ws.merge_cells('A1:G1')

        row = 3
        headers = ['Milestone ID', 'Milestone', 'Description', 'Planned Date', 'Actual Date', 'Owner', 'Status']
        for col, header in enumerate(headers, 1):
            ExcelFormatter.format_header_row(ws, row, [col])
            ws.cell(row=row, column=col).value = header

        row += 1
        plan_data = ProjectPlanEngine.generate_project_plan(project_info)

        for idx, milestone in enumerate(plan_data.get('milestones', []), 1):
            ExcelFormatter.format_data_row(ws, row, list(range(1, 8)))
            ws.cell(row=row, column=1).value = f"M-{idx:03d}"
            ws.cell(row=row, column=2).value = milestone.get('milestone', '')
            ws.cell(row=row, column=3).value = ''
            ws.cell(row=row, column=4).value = milestone.get('date', '')
            ws.cell(row=row, column=5).value = ''
            ws.cell(row=row, column=6).value = 'PM'
            ws.cell(row=row, column=7).value = milestone.get('status', 'Planned')

            # Status color coding
            status = milestone.get('status', 'Planned')
            if status == 'Completed':
                color = WorkbookEnhancements.COLOR_PALETTE['success']
            elif status == 'In Progress':
                color = WorkbookEnhancements.COLOR_PALETTE['warning']
            else:
                color = WorkbookEnhancements.COLOR_PALETTE['primary_light']

            ws.cell(row=row, column=7).fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            row += 1

        ExcelFormatter.set_column_widths(ws, {'A': 12, 'B': 20, 'C': 30, 'D': 12, 'E': 12, 'F': 12, 'G': 12})
        ExcelFormatter.freeze_panes(ws, 3)

    @staticmethod
    def create_resource_plan(workbook, project_info: Dict, db_summary: Dict):
        """Create Resource Planning sheet."""
        ws = workbook.create_sheet('04_Resource_Plan')

        ExcelFormatter.format_title(ws, 1, 1, 'Resource Plan')
        ws.merge_cells('A1:H1')

        row = 3
        headers = ['Resource Name', 'Role', 'Department', 'Allocation %', 'Start Date', 'End Date', 'Effort Hours', 'Cost Estimate']
        for col, header in enumerate(headers, 1):
            ExcelFormatter.format_header_row(ws, row, [col])
            ws.cell(row=row, column=col).value = header

        row += 1
        team_members = db_summary.get('team_members', [])
        start_date = project_info.get('start_date', 'TBD')
        duration_weeks = project_info.get('duration_weeks', 0)

        for member in team_members:
            ExcelFormatter.format_data_row(ws, row, list(range(1, 9)))
            ws.cell(row=row, column=1).value = f"{member.get('role', '')} Team"
            ws.cell(row=row, column=2).value = member.get('role', '')
            ws.cell(row=row, column=3).value = 'Project'
            ws.cell(row=row, column=4).value = 1.0
            ExcelFormatter.format_percentage_column(ws, row, row, 4)
            ws.cell(row=row, column=5).value = start_date
            ws.cell(row=row, column=6).value = start_date
            ws.cell(row=row, column=7).value = duration_weeks * 40 * member.get('count', 1)
            ws.cell(row=row, column=8).value = duration_weeks * 40 * 100 * member.get('count', 1)
            row += 1

        ExcelFormatter.set_column_widths(ws, {'A': 18, 'B': 15, 'C': 15, 'D': 12, 'E': 12, 'F': 12, 'G': 12, 'H': 15})
        ExcelFormatter.freeze_panes(ws, 3)

    @staticmethod
    def create_raid_register(workbook, project_info: Dict, db_summary: Dict):
        """Create unified RAID Register (Risk, Assumption, Issue, Dependency)."""
        ws = workbook.create_sheet('05_RAID_Register')

        ExcelFormatter.format_title(ws, 1, 1, 'RAID Register')
        ws.merge_cells('A1:J1')

        row = 3
        headers = ['ID', 'Type', 'Description', 'Impact', 'Probability/Status', 'Owner', 'Mitigation/Action', 'Target Date', 'Closure Date', 'Status']
        for col, header in enumerate(headers, 1):
            ExcelFormatter.format_header_row(ws, row, [col])
            ws.cell(row=row, column=col).value = header

        row += 1
        risks = db_summary.get('risks', [])

        for idx, risk in enumerate(risks, 1):
            ExcelFormatter.format_data_row(ws, row, list(range(1, 11)))
            ws.cell(row=row, column=1).value = f"R-{idx:03d}"
            ws.cell(row=row, column=2).value = 'Risk'
            ws.cell(row=row, column=3).value = risk.get('risk_description', '')
            ws.cell(row=row, column=4).value = risk.get('severity', 'Medium')
            ws.cell(row=row, column=5).value = 0.5
            ExcelFormatter.format_percentage_column(ws, row, row, 5)
            ws.cell(row=row, column=6).value = 'PM'
            ws.cell(row=row, column=7).value = risk.get('mitigation', '')
            ws.cell(row=row, column=8).value = ''
            ws.cell(row=row, column=9).value = ''
            ws.cell(row=row, column=10).value = 'Active'

            # Type color coding
            type_colors = {'Risk': 'F4B084', 'Assumption': 'FFE699', 'Issue': 'F4B084', 'Dependency': 'BFBFBF'}
            ws.cell(row=row, column=2).fill = PatternFill(start_color=type_colors.get('Risk', 'F4B084'),
                                                          end_color=type_colors.get('Risk', 'F4B084'),
                                                          fill_type='solid')
            row += 1

        ExcelFormatter.set_column_widths(ws, {
            'A': 10, 'B': 12, 'C': 25, 'D': 12, 'E': 12, 'F': 12, 'G': 20, 'H': 12, 'I': 12, 'J': 12
        })
        ExcelFormatter.freeze_panes(ws, 3)

    @staticmethod
    def create_leave_capacity_planner(workbook, project_info: Dict, db_summary: Dict):
        """Create Leave and Capacity Planner."""
        ws = workbook.create_sheet('06_Leave_Capacity_Planner')

        ExcelFormatter.format_title(ws, 1, 1, 'Leave & Capacity Planner')
        ws.merge_cells('A1:G1')

        row = 3
        headers = ['Resource', 'Role', 'Allocation %', 'Leave Days', 'Available Capacity %', 'Utilization %', 'Remarks']
        for col, header in enumerate(headers, 1):
            ExcelFormatter.format_header_row(ws, row, [col])
            ws.cell(row=row, column=col).value = header

        row += 1
        team_members = db_summary.get('team_members', [])

        for member in team_members:
            ExcelFormatter.format_data_row(ws, row, list(range(1, 8)))
            ws.cell(row=row, column=1).value = f"{member.get('role', '')} Team"
            ws.cell(row=row, column=2).value = member.get('role', '')
            ws.cell(row=row, column=3).value = 1.0
            ExcelFormatter.format_percentage_column(ws, row, row, 3)
            ws.cell(row=row, column=4).value = 0
            ws.cell(row=row, column=5).value = 1.0
            ExcelFormatter.format_percentage_column(ws, row, row, 5)
            ws.cell(row=row, column=6).value = 0.8
            ExcelFormatter.format_percentage_column(ws, row, row, 6)
            ws.cell(row=row, column=7).value = ''
            row += 1

        ExcelFormatter.set_column_widths(ws, {'A': 18, 'B': 15, 'C': 12, 'D': 12, 'E': 15, 'F': 12, 'G': 20})
        ExcelFormatter.freeze_panes(ws, 3)

    @staticmethod
    def create_weekly_status_tracker(workbook, project_info: Dict):
        """Create Weekly Status Tracker."""
        ws = workbook.create_sheet('07_Weekly_Status')

        ExcelFormatter.format_title(ws, 1, 1, 'Weekly Project Status')
        ws.merge_cells('A1:E1')

        row = 3
        headers = ['Week', 'Planned %', 'Actual %', 'Variance', 'Status']
        for col, header in enumerate(headers, 1):
            ExcelFormatter.format_header_row(ws, row, [col])
            ws.cell(row=row, column=col).value = header

        row += 1
        duration_weeks = project_info.get('duration_weeks', 0)

        for week in range(1, int(duration_weeks) + 1):
            ExcelFormatter.format_data_row(ws, row, list(range(1, 6)))
            ws.cell(row=row, column=1).value = f"Week {week}"
            ws.cell(row=row, column=2).value = week / duration_weeks
            ExcelFormatter.format_percentage_column(ws, row, row, 2)
            ws.cell(row=row, column=3).value = 0
            ExcelFormatter.format_percentage_column(ws, row, row, 3)
            ws.cell(row=row, column=4).value = 0
            ExcelFormatter.format_percentage_column(ws, row, row, 4)
            ws.cell(row=row, column=5).value = 'Planned'
            row += 1

        ExcelFormatter.set_column_widths(ws, {'A': 15, 'B': 12, 'C': 12, 'D': 12, 'E': 15})
        ExcelFormatter.freeze_panes(ws, 3)

    @staticmethod
    def create_ai_project_summary(workbook, project_info: Dict, db_summary: Dict):
        """Create AI-generated Project Summary sheet."""
        ws = workbook.create_sheet('08_AI_Project_Summary')

        ExcelFormatter.format_title(ws, 1, 1, 'AI Project Summary')
        ws.merge_cells('A1:D1')

        row = 3

        # Generate AI insights
        duration = project_info.get('duration_weeks', 16)
        team_size = project_info.get('team_size', 5)
        effort_hours = duration * 40 * team_size
        project_type = project_info.get('project_type', 'Unknown')
        risks = len(db_summary.get('risks', []))

        summary_sections = [
            ('PROJECT OVERVIEW', [
                f"Project Type: {project_type}",
                f"Expected Duration: {duration} weeks",
                f"Team Size: {team_size} resources",
                f"Estimated Effort: {effort_hours:,} hours"
            ]),
            ('KEY INSIGHTS', [
                f"Total Project Risks: {risks}",
                f"Resource Utilization: 80%",
                "Critical Path: Identified",
                "Recommended Delivery Model: Agile" if 'Agile' in project_type else "Recommended Delivery Model: Waterfall"
            ]),
            ('RECOMMENDED GOVERNANCE', [
                "Weekly Status Reviews",
                "Bi-weekly Risk Reviews",
                "Monthly Steering Committee Meetings",
                "Real-time Dashboard Monitoring"
            ]),
            ('SUCCESS CRITERIA', [
                "On-time delivery within schedule",
                "Budget alignment with estimates",
                "Quality metrics met",
                "Stakeholder satisfaction > 90%"
            ])
        ]

        for section_title, items in summary_sections:
            ws.merge_cells(f'A{row}:D{row}')
            section_cell = ws[f'A{row}']
            section_cell.value = section_title
            section_cell.font = Font(bold=True, size=12, color=WorkbookEnhancements.COLOR_PALETTE['white'])
            section_cell.fill = PatternFill(start_color=WorkbookEnhancements.COLOR_PALETTE['primary'],
                                           end_color=WorkbookEnhancements.COLOR_PALETTE['primary'],
                                           fill_type='solid')
            row += 1

            for item in items:
                ws.merge_cells(f'A{row}:D{row}')
                item_cell = ws[f'A{row}']
                item_cell.value = f"• {item}"
                item_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                row += 1

            row += 1

        ExcelFormatter.set_column_widths(ws, {'A': 40, 'B': 20, 'C': 20, 'D': 20})
        ExcelFormatter.freeze_panes(ws, 3)

    @staticmethod
    def _calculate_end_date(project_info: Dict) -> str:
        """Calculate project end date."""
        try:
            start = datetime.strptime(project_info.get('start_date', '2025-01-01'), '%Y-%m-%d')
            weeks = int(project_info.get('duration_weeks', 0))
            end = start + timedelta(weeks=weeks)
            return end.strftime('%Y-%m-%d')
        except:
            return 'TBD'

    @staticmethod
    def _calculate_project_health(project_info: Dict, db_summary: Dict) -> Dict:
        """Calculate project health status."""
        risk_count = len(db_summary.get('risks', []))
        complexity_score = len(db_summary.get('team_members', [])) + risk_count

        if risk_count > 5 or complexity_score > 15:
            return {'status': 'RED - High Risk', 'color': 'C5504F'}
        elif risk_count > 2 or complexity_score > 8:
            return {'status': 'AMBER - Medium Risk', 'color': 'FFC000'}
        else:
            return {'status': 'GREEN - Low Risk', 'color': '70AD47'}
