"""
Dynamic Gantt Chart Generator for Project Aura.
Generates professional Gantt charts with timeline visualization.
"""

import logging
from datetime import datetime, timedelta
from typing import Dict, Any, List
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class GanttGenerator:
    """Generates professional Gantt charts in Excel."""

    COLOR_PALETTE = {
        'complete': '70AD47',
        'in_progress': 'FFC000',
        'planned': 'BFBFBF',
        'critical': 'C5504F',
        'header': '366092'
    }

    @staticmethod
    def create_gantt_chart(workbook, project_info: Dict, plan_data: Dict):
        """
        Create dynamic Gantt chart sheet.

        Args:
            workbook: Excel workbook
            project_info: Project information
            plan_data: Plan data with phases and timeline
        """
        ws = workbook.create_sheet('09_Gantt_Chart')

        # Title
        ws.merge_cells('A1:L1')
        title = ws['A1']
        title.value = f"{project_info.get('client_name', 'Client')} - Project Timeline Gantt Chart"
        title.font = {'bold': True, 'size': 14, 'color': 'FFFFFF'}
        title.fill = PatternFill(start_color=GanttGenerator.COLOR_PALETTE['header'],
                               end_color=GanttGenerator.COLOR_PALETTE['header'],
                               fill_type='solid')
        title.alignment = Alignment(horizontal='center', vertical='center')

        # Parse dates
        try:
            start_date = datetime.strptime(project_info.get('start_date', '2025-01-01'), '%Y-%m-%d')
        except:
            start_date = datetime(2025, 1, 1)

        duration_weeks = int(project_info.get('duration_weeks', 16))
        project_duration_days = duration_weeks * 7

        # Create header row with week numbers
        row = 3
        ws['A3'] = 'Phase/Task'
        ws['B3'] = 'Start'
        ws['C3'] = 'End'
        ws['D3'] = 'Days'

        # Date headers
        col = 5
        for week in range(duration_weeks + 1):
            date = start_date + timedelta(weeks=week)
            header_cell = ws.cell(row=3, column=col)
            header_cell.value = f"W{week+1}"
            header_cell.font = {'bold': True, 'size': 9}
            header_cell.fill = PatternFill(start_color=GanttGenerator.COLOR_PALETTE['header'],
                                          end_color=GanttGenerator.COLOR_PALETTE['header'],
                                          fill_type='solid')
            header_cell.alignment = Alignment(horizontal='center', vertical='center')
            header_cell.font = {'color': 'FFFFFF', 'bold': True, 'size': 9}
            col += 1

        # Add phases
        row = 4
        phases = plan_data.get('phases', [])

        current_date = start_date
        for phase in phases:
            phase_name = phase.get('phase', 'Phase')
            phase_days = phase.get('duration_days', 7)
            phase_end = current_date + timedelta(days=phase_days)

            # Phase row
            ws.cell(row=row, column=1).value = phase_name
            ws.cell(row=row, column=2).value = current_date.strftime('%Y-%m-%d')
            ws.cell(row=row, column=3).value = phase_end.strftime('%Y-%m-%d')
            ws.cell(row=row, column=4).value = phase_days

            # Calculate which columns to color (timeline visualization)
            start_week = (current_date - start_date).days // 7
            end_week = (phase_end - start_date).days // 7

            for col in range(5, 5 + duration_weeks + 1):
                week_num = col - 5
                cell = ws.cell(row=row, column=col)

                if week_num >= start_week and week_num <= end_week:
                    # Determine color
                    if phase.get('status') == 'Completed':
                        color = GanttGenerator.COLOR_PALETTE['complete']
                    elif phase.get('status') == 'In Progress':
                        color = GanttGenerator.COLOR_PALETTE['in_progress']
                    else:
                        color = GanttGenerator.COLOR_PALETTE['planned']

                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    cell.value = '█'
                    cell.font = {'size': 8, 'bold': True}
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            # Format date columns
            ws.cell(row=row, column=2).number_format = 'YYYY-MM-DD'
            ws.cell(row=row, column=3).number_format = 'YYYY-MM-DD'

            current_date = phase_end + timedelta(days=1)
            row += 1

        # Legend
        row += 2
        legend_data = [
            ('█ Planned', GanttGenerator.COLOR_PALETTE['planned']),
            ('█ In Progress', GanttGenerator.COLOR_PALETTE['in_progress']),
            ('█ Completed', GanttGenerator.COLOR_PALETTE['complete']),
        ]

        for text, color in legend_data:
            ws.cell(row=row, column=1).value = text
            ws.cell(row=row, column=1).fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            ws.cell(row=row, column=1).font = {'bold': True, 'size': 10}
            row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 8

        for col in range(5, 5 + duration_weeks + 1):
            ws.column_dimensions[get_column_letter(col)].width = 4

        # Freeze panes
        ws.freeze_panes = 'E4'

    @staticmethod
    def create_timeline_visualization(workbook, project_info: Dict, plan_data: Dict):
        """Create simplified timeline visualization sheet."""
        ws = workbook.create_sheet('10_Timeline')

        ws.merge_cells('A1:H1')
        title = ws['A1']
        title.value = 'Project Timeline Summary'
        title.font = {'bold': True, 'size': 14, 'color': 'FFFFFF'}
        title.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

        row = 3
        ws.cell(row=row, column=1).value = 'Phase'
        ws.cell(row=row, column=2).value = 'Duration'
        ws.cell(row=row, column=3).value = 'Progress'
        ws.cell(row=row, column=4).value = 'Status'

        phases = plan_data.get('phases', [])
        row += 1

        for phase in phases:
            ws.cell(row=row, column=1).value = phase.get('phase', '')
            ws.cell(row=row, column=2).value = f"{phase.get('duration_days', 0)} days"
            ws.cell(row=row, column=3).value = 0  # Progress percentage
            ws.cell(row=row, column=4).value = phase.get('status', 'Planned')

            # Status color
            status = phase.get('status', 'Planned')
            if status == 'Completed':
                color = '70AD47'
            elif status == 'In Progress':
                color = 'FFC000'
            else:
                color = 'BFBFBF'

            ws.cell(row=row, column=4).fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            row += 1

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
