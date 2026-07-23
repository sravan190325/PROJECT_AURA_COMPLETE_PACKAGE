"""
Sample-format workbook generator for Project Aura.

Uses ABC_Retail_Project_Plan_Improved.xlsx as the workbook template so
downloaded project plans keep the same sheet order, layout, formulas,
dropdowns, widths, frozen panes, and visual formatting as the approved sample.
"""

import logging
import os
from copy import copy
from datetime import datetime, timedelta
from typing import Any, Dict, List

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from services.project_plan_engine import ProjectPlanEngine

logger = logging.getLogger(__name__)


class SampleFormatWorkbookGenerator:
    """Generate workbooks in the ABC_Retail_Project_Plan_Improved format."""

    TEMPLATE_FILENAME = "ABC_Retail_Project_Plan_Improved.xlsx"
    STATUS_OPTIONS = "Not Started,In Progress,Done,On Hold,Cancelled"
    MILESTONE_STATUS_OPTIONS = "Planned,In Progress,Complete,Delayed,Cancelled"
    CHANGE_TYPE_OPTIONS = "Scope,Schedule,Budget,Resource,Other"
    CHANGE_STATUS_OPTIONS = "Raised,Under Review,Approved,Rejected,Deferred"

    ROLE_DEPARTMENTS = {
        "PM": "Delivery",
        "Project Manager": "Delivery",
        "BA": "Business Analysis",
        "Business Analyst": "Business Analysis",
        "Architect": "Architecture",
        "Tech Lead": "Engineering",
        "Developer": "Engineering",
        "QA Engineer": "Quality",
        "QA Lead": "Quality",
        "DevOps": "Engineering",
        "Client PM": "Client",
    }

    ROLE_DAILY_RATES = {
        "PM": 650,
        "Project Manager": 650,
        "BA": 550,
        "Business Analyst": 550,
        "Architect": 800,
        "Tech Lead": 750,
        "Developer": 600,
        "QA Engineer": 500,
        "QA Lead": 600,
        "DevOps": 650,
        "Client PM": 0,
    }

    DEFAULT_RACI_ACTIVITIES = [
        "Project Kickoff",
        "Requirements Gathering",
        "Solution Architecture",
        "Project Plan Approval",
        "Development",
        "Testing",
        "UAT Sign-off",
        "Go-Live",
        "Hypercare",
        "Project Closure",
    ]

    def __init__(self, project_info: Dict[str, Any], db_summary: Dict[str, Any]):
        self.project_info = project_info or {}
        self.db_summary = db_summary or {}
        self.workbook = None

    def generate(self, output_path: str) -> bool:
        """Generate the workbook and save it to output_path."""
        try:
            template_path = self._template_path()
            if not os.path.exists(template_path):
                logger.error("Workbook template not found: %s", template_path)
                return False

            self.workbook = load_workbook(template_path)
            context = self._build_context()

            self._populate_home(context)
            self._populate_executive_dashboard(context)
            self._populate_task_plan(context)
            self._populate_budget_tracker(context)
            self._populate_milestone_tracker(context)
            self._populate_resource_plan(context)
            self._populate_raid_register(context)
            self._populate_raci_matrix(context)
            self._populate_change_log(context)
            self._set_properties(context)

            os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
            self.workbook.save(output_path)
            logger.info("Sample-format workbook generated successfully: %s", output_path)
            return True
        except Exception as exc:
            logger.error("Error generating sample-format workbook: %s", exc, exc_info=True)
            return False

    def _template_path(self) -> str:
        """Return the absolute template path."""
        project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
        return os.path.join(project_root, self.TEMPLATE_FILENAME)

    def _build_context(self) -> Dict[str, Any]:
        start = self._parse_date(self.project_info.get("start_date")) or datetime.now()
        duration_weeks = self._safe_int(self.project_info.get("duration_weeks"), 12)
        end = start + timedelta(weeks=duration_weeks)

        plan = ProjectPlanEngine.generate_project_plan(
            {
                **self.project_info,
                "start_date": start.strftime("%Y-%m-%d"),
                "duration_weeks": duration_weeks,
                "team_size": self._safe_int(self.project_info.get("team_size"), 1),
            }
        )
        phases = plan.get("phases") if plan.get("success") else []
        if not phases:
            phases = self._fallback_phases(start, duration_weeks)

        team = self._normalized_team(start, end)
        tasks = self._build_tasks(phases, team)
        milestones = self._build_milestones(plan.get("milestones", []), phases)
        risks = self._build_risks()
        assumptions = self._build_assumptions()
        dependencies = self._build_dependencies(plan.get("dependencies", []), phases)

        return {
            "client": self.project_info.get("client_name") or "Client",
            "project": self.project_info.get("project_name") or "Project",
            "project_type": self.project_info.get("project_type") or "Project",
            "delivery_model": self.project_info.get("delivery_model") or "TBD",
            "scope": self.project_info.get("scope") or "TBD",
            "start": start,
            "end": end,
            "duration_weeks": duration_weeks,
            "phases": phases,
            "team": team,
            "tasks": tasks,
            "milestones": milestones,
            "risks": risks,
            "assumptions": assumptions,
            "dependencies": dependencies,
            "deliverables": self.db_summary.get("deliverables", []),
        }

    def _populate_home(self, context: Dict[str, Any]) -> None:
        ws = self.workbook["00_Home"]
        ws["A1"] = "PROJECT AURA - AI-Powered Project Management"
        ws["A2"] = f"{context['client']} - {context['project']}"
        ws["B5"] = context["project"]
        ws["B6"] = context["client"]
        ws["B7"] = "Project Aura"
        ws["B8"] = context["project_type"]
        ws["B9"] = context["delivery_model"]
        ws["B10"] = context["start"]
        ws["B11"] = context["end"]
        ws["B12"] = f"{context['duration_weeks']} weeks"
        ws["B13"] = len(context["team"])
        ws["B14"] = datetime.now().strftime("%Y-%m-%d %H:%M")

        open_risks = len(context["risks"])
        ws["F5"] = self._rag_message(open_risks)
        ws["F6"] = "AMBER - update actuals in Task Plan"
        ws["F7"] = "GREEN - update in Budget Tracker"
        ws["F8"] = self._resource_message(context["team"])
        ws["F9"] = f"{'RED' if open_risks >= 5 else 'AMBER' if open_risks else 'GREEN'} - {open_risks} open risks"

        scope = context["scope"]
        ws["B19"] = scope
        ws["B20"] = "Review task plan, RAID register, budget tracker, and change log weekly."
        ws["B24"] = "Blue cells are inputs. Green cells are formulas. Keep change approvals in Change_Log."
        ws["B26"] = f"Generated by Project Aura on {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        self._format_dates(ws, ["B10", "B11"])

    def _populate_executive_dashboard(self, context: Dict[str, Any]) -> None:
        ws = self.workbook["Executive_Dashboard"]
        self._clear_values(ws, 9, 17, 1, 10)

        ws["A5"] = '=IFERROR(COUNTIF(Detailed_Task_Plan!M4:M200,"Done")/COUNTIF(Detailed_Task_Plan!M4:M200,"<>"),0)'
        ws["C5"] = '=IFERROR(Budget_Tracker!G14/Budget_Tracker!E14,0)'
        ws["E5"] = '=COUNTIF(Milestone_Tracker!F3:F40,"Complete")'
        ws["G5"] = '=COUNTIF(RAID_Register!I5:I50,"Open")'
        ws["I5"] = '=COUNTIF(RAID_Register!I54:I100,"Open")'

        for row_idx, phase in enumerate(context["phases"], 9):
            self._copy_row_style(ws, 9, row_idx, 1, 10)
            ws.cell(row_idx, 1, phase.get("phase"))
            ws.cell(row_idx, 2, phase.get("start_date"))
            ws.cell(row_idx, 3, phase.get("end_date"))
            ws.cell(row_idx, 4, f"{phase.get('duration_weeks', 1)}w")
            ws.cell(row_idx, 5, self._owner_for_phase(phase.get("phase", "")))
            ws.cell(row_idx, 6, phase.get("status", "Planned"))
            ws.cell(row_idx, 7, "0%")
            ws.cell(row_idx, 8, "AMBER")

        ws["A19"] = "Current Focus"
        ws["B19"] = context["phases"][0].get("phase", "Initiation") if context["phases"] else "Initiation"
        ws["A20"] = "Next Milestone"
        ws["B20"] = context["milestones"][0]["milestone"] if context["milestones"] else "Project Kickoff"
        ws["A21"] = "Top Delivery Risk"
        ws["B21"] = context["risks"][0]["description"] if context["risks"] else "No major risks identified"
        ws["A22"] = "Executive Action"
        ws["B22"] = "Review RAID and Change_Log before approving baseline changes."

    def _populate_task_plan(self, context: Dict[str, Any]) -> None:
        ws = self.workbook["Detailed_Task_Plan"]
        self._clear_values(ws, 4, max(ws.max_row, 220), 1, 14)

        for row_idx, task in enumerate(context["tasks"], 4):
            self._copy_row_style(ws, 4, row_idx, 1, 14)
            ws.cell(row_idx, 1, task["wbs"])
            ws.cell(row_idx, 2, task["phase"])
            ws.cell(row_idx, 3, task["task"])
            ws.cell(row_idx, 4, task["owner"])
            ws.cell(row_idx, 5, task.get("predecessor"))
            ws.cell(row_idx, 6, task["start"])
            ws.cell(row_idx, 7, task["end"])
            ws.cell(row_idx, 8, None)
            ws.cell(row_idx, 9, None)
            ws.cell(row_idx, 10, f'=IFERROR(DATEDIF(F{row_idx},G{row_idx},"D"),"")')
            ws.cell(row_idx, 11, f'=IFERROR(DATEDIF(H{row_idx},I{row_idx},"D"),"")')
            ws.cell(row_idx, 12, "0%")
            ws.cell(row_idx, 13, "Not Started")
            ws.cell(row_idx, 14, None)

        self._replace_list_validation(ws, "M4:M200", self.STATUS_OPTIONS)

    def _populate_budget_tracker(self, context: Dict[str, Any]) -> None:
        ws = self.workbook["Budget_Tracker"]
        self._clear_values(ws, 10, max(ws.max_row, 40), 1, 9)

        contract_value = max(len(context["team"]) * context["duration_weeks"] * 5 * 600, 50000)
        ws["B4"] = contract_value
        ws["B5"] = round(contract_value * 0.10, 2)
        ws["B6"] = "=B4-B5"

        for row_idx, member in enumerate(context["team"], 10):
            self._copy_row_style(ws, 10, row_idx, 1, 9)
            role = member["role"]
            ws.cell(row_idx, 1, member["resource"])
            ws.cell(row_idx, 2, role)
            ws.cell(row_idx, 3, self.ROLE_DAILY_RATES.get(role, 600))
            ws.cell(row_idx, 4, max(1, int(context["duration_weeks"] * 5 * 0.8)))
            ws.cell(row_idx, 5, f"=C{row_idx}*D{row_idx}")
            ws.cell(row_idx, 6, 0)
            ws.cell(row_idx, 7, f"=C{row_idx}*F{row_idx}")
            ws.cell(row_idx, 8, f"=E{row_idx}-G{row_idx}")
            ws.cell(row_idx, 9, f"=IFERROR(H{row_idx}/E{row_idx},0)")

        total_row = max(14, 10 + len(context["team"]))
        self._copy_row_style(ws, 14, total_row, 1, 9)
        ws.cell(total_row, 1, "TOTAL")
        ws.cell(total_row, 5, f"=SUM(E10:E{total_row - 1})")
        ws.cell(total_row, 7, f"=SUM(G10:G{total_row - 1})")
        ws.cell(total_row, 8, f"=SUM(H10:H{total_row - 1})")
        ws.cell(total_row, 9, f"=IFERROR(H{total_row}/E{total_row},0)")

    def _populate_milestone_tracker(self, context: Dict[str, Any]) -> None:
        ws = self.workbook["Milestone_Tracker"]
        self._clear_values(ws, 3, max(ws.max_row, 50), 1, 9)

        for row_idx, milestone in enumerate(context["milestones"], 3):
            self._copy_row_style(ws, 3, row_idx, 1, 9)
            ws.cell(row_idx, 1, f"M{row_idx - 2}")
            ws.cell(row_idx, 2, milestone["milestone"])
            ws.cell(row_idx, 3, self._owner_for_phase(milestone.get("phase", "")))
            ws.cell(row_idx, 4, milestone["date"])
            ws.cell(row_idx, 5, None)
            ws.cell(row_idx, 6, milestone.get("status", "Planned"))
            ws.cell(row_idx, 7, milestone.get("priority", "High"))
            ws.cell(row_idx, 8, f'=IFERROR(IF(E{row_idx}="","",DATEDIF(D{row_idx},E{row_idx},"D")),"")')
            ws.cell(row_idx, 9, None)

        self._replace_list_validation(ws, "F3:F50", self.MILESTONE_STATUS_OPTIONS)

    def _populate_resource_plan(self, context: Dict[str, Any]) -> None:
        ws = self.workbook["Resource_Plan"]
        self._clear_values(ws, 3, max(ws.max_row, 40), 1, 10)

        for row_idx, member in enumerate(context["team"], 3):
            self._copy_row_style(ws, 3, row_idx, 1, 10)
            role = member["role"]
            ws.cell(row_idx, 1, member["resource"])
            ws.cell(row_idx, 2, role)
            ws.cell(row_idx, 3, self.ROLE_DEPARTMENTS.get(role, "Delivery"))
            ws.cell(row_idx, 4, context["start"].strftime("%Y-%m-%d"))
            ws.cell(row_idx, 5, context["end"].strftime("%Y-%m-%d"))
            ws.cell(row_idx, 6, int(context["duration_weeks"] * 40 * 0.8))
            ws.cell(row_idx, 7, 0)
            ws.cell(row_idx, 8, "80%")
            ws.cell(row_idx, 9, f"=IFERROR(G{row_idx}/F{row_idx},0)")
            ws.cell(row_idx, 10, f'=IF(IFERROR(VALUE(LEFT(H{row_idx},LEN(H{row_idx})-1)),0)>90,"OVERLOADED","OK")')

        note_row = max(9, 4 + len(context["team"]))
        self._copy_row_style(ws, 9, note_row, 1, 10)
        ws.cell(note_row, 1, "Note: Max recommended allocation is 80% to allow for meetings, admin, and leave. Flag highlights resources >90%.")

    def _populate_raid_register(self, context: Dict[str, Any]) -> None:
        ws = self.workbook["RAID_Register"]
        self._clear_values(ws, 5, 10, 1, 9)
        self._clear_values(ws, 14, 18, 1, 9)
        self._clear_values(ws, 22, 22, 1, 9)
        self._clear_values(ws, 26, 50, 1, 9)

        for row_idx, risk in enumerate(context["risks"][:20], 5):
            self._copy_row_style(ws, 5, row_idx, 1, 9)
            ws.cell(row_idx, 1, f"R{row_idx - 4}")
            ws.cell(row_idx, 2, "Risk")
            ws.cell(row_idx, 3, risk["description"])
            ws.cell(row_idx, 4, risk["impact"])
            ws.cell(row_idx, 5, risk["probability"])
            ws.cell(row_idx, 6, f'=IF(D{row_idx}="","",D{row_idx}*E{row_idx})')
            ws.cell(row_idx, 7, risk["owner"])
            ws.cell(row_idx, 8, risk["mitigation"])
            ws.cell(row_idx, 9, risk["status"])

        for offset, assumption in enumerate(context["assumptions"][:5], 14):
            self._copy_row_style(ws, 14, offset, 1, 9)
            ws.cell(offset, 1, f"A{offset - 13}")
            ws.cell(offset, 2, "Assumption")
            ws.cell(offset, 3, assumption)
            ws.cell(offset, 7, "PM")
            ws.cell(offset, 8, "Validate during planning and weekly governance.")
            ws.cell(offset, 9, "Open")

        ws["A22"] = "I1"
        ws["B22"] = "Issue"
        ws["C22"] = "No open issues logged at generation time."
        ws["G22"] = "PM"
        ws["H22"] = "Track new issues here as they arise."
        ws["I22"] = "Open"

        for row_idx, dep in enumerate(context["dependencies"][:15], 26):
            self._copy_row_style(ws, 26, row_idx, 1, 9)
            ws.cell(row_idx, 1, f"D{row_idx - 25}")
            ws.cell(row_idx, 2, "Dependency")
            ws.cell(row_idx, 3, f"{dep.get('task', 'Task')} depends on {dep.get('depends_on', 'prior work')}"
                    )
            ws.cell(row_idx, 7, "PM")
            ws.cell(row_idx, 8, dep.get("dependency_type", "Finish to Start"))
            ws.cell(row_idx, 9, "Open")

        self._replace_list_validation(ws, "I5:I50 I54:I100", "Open,In Progress,Closed,Accepted")

    def _populate_raci_matrix(self, context: Dict[str, Any]) -> None:
        ws = self.workbook["RACI_Matrix"]
        self._clear_values(ws, 5, max(ws.max_row, 50), 1, 10)

        roles = [member["role"] for member in context["team"][:7]]
        while len(roles) < 7:
            roles.append(["PM", "BA", "Architect", "Dev Lead", "QA Lead", "DevOps", "Client PM"][len(roles)])

        for col, role in enumerate(roles, 2):
            ws.cell(4, col, role)

        activities = [task["task"] for task in context["tasks"][:18]] or self.DEFAULT_RACI_ACTIVITIES
        for row_idx, activity in enumerate(activities, 5):
            self._copy_row_style(ws, 5, row_idx, 1, 10)
            ws.cell(row_idx, 1, activity)
            for col_idx, role in enumerate(roles, 2):
                ws.cell(row_idx, col_idx, self._raci_value(activity, role, col_idx))

    def _populate_change_log(self, context: Dict[str, Any]) -> None:
        ws = self.workbook["Change_Log"]
        self._clear_values(ws, 4, max(ws.max_row, 80), 1, 11)
        ws["A4"] = "CR-001"
        ws["B4"] = context["start"].strftime("%Y-%m-%d")
        ws["C4"] = "PM"
        ws["D4"] = "Scope"
        ws["E4"] = "[Example] Add approved project scope change here"
        ws["F4"] = "0 days"
        ws["G4"] = "$0"
        ws["H4"] = "Sponsor"
        ws["I4"] = context["start"].strftime("%Y-%m-%d")
        ws["J4"] = "Raised"
        ws["K4"] = context["start"].strftime("%Y-%m-%d")
        self._replace_list_validation(ws, "D4:D80", self.CHANGE_TYPE_OPTIONS)
        self._replace_list_validation(ws, "J4:J80", self.CHANGE_STATUS_OPTIONS)

    def _set_properties(self, context: Dict[str, Any]) -> None:
        self.workbook.properties.title = f"{context['client']} - {context['project']} Project Plan"
        self.workbook.properties.subject = f"{context['project_type']} Project Planning"
        self.workbook.properties.creator = "Project Aura"
        self.workbook.properties.company = context["client"]
        self.workbook.properties.modified = datetime.now()

    def _normalized_team(self, start: datetime, end: datetime) -> List[Dict[str, str]]:
        members = self.db_summary.get("team_members", [])
        team: List[Dict[str, str]] = []
        for member in members:
            role = member.get("role") or "Team Member"
            count = max(1, self._safe_int(member.get("count"), 1))
            for index in range(1, count + 1):
                resource = f"{role.replace(' ', '')}_{index}"
                team.append({"role": role, "resource": resource})

        if team:
            return team

        defaults = ["PM", "BA", "Architect", "Developer", "QA Engineer", "DevOps"]
        team_size = max(1, self._safe_int(self.project_info.get("team_size"), len(defaults)))
        return [
            {"role": defaults[index % len(defaults)], "resource": f"{defaults[index % len(defaults)].replace(' ', '')}_{index + 1}"}
            for index in range(team_size)
        ]

    def _build_tasks(self, phases: List[Dict[str, Any]], team: List[Dict[str, str]]) -> List[Dict[str, Any]]:
        tasks: List[Dict[str, Any]] = []
        for phase_index, phase in enumerate(phases, 1):
            phase_name = phase.get("phase", f"Phase {phase_index}")
            start = phase.get("start_date")
            end = phase.get("end_date")
            owner = self._owner_for_phase(phase_name)
            predecessor = f"{phase_index - 1}.0" if phase_index > 1 else None
            tasks.append(
                {
                    "wbs": f"{phase_index}.0",
                    "phase": phase_name,
                    "task": f"{phase_name} complete",
                    "owner": owner,
                    "predecessor": predecessor,
                    "start": start,
                    "end": end,
                }
            )
            mid = self._midpoint_date(start, end)
            tasks.append(
                {
                    "wbs": f"{phase_index}.1",
                    "phase": phase_name,
                    "task": f"{phase_name} planning and preparation",
                    "owner": owner,
                    "predecessor": predecessor,
                    "start": start,
                    "end": mid,
                }
            )
            tasks.append(
                {
                    "wbs": f"{phase_index}.2",
                    "phase": phase_name,
                    "task": f"{phase_name} execution and sign-off",
                    "owner": owner,
                    "predecessor": f"{phase_index}.1",
                    "start": mid,
                    "end": end,
                }
            )

        for deliverable in self.db_summary.get("deliverables", []):
            if len(tasks) >= 190:
                break
            phase_index = min(len(phases), max(1, len(tasks) // 3))
            phase = phases[phase_index - 1]
            tasks.append(
                {
                    "wbs": f"{phase_index}.{len(tasks)}",
                    "phase": phase.get("phase"),
                    "task": deliverable.get("deliverable_name") or "Deliverable",
                    "owner": "PM",
                    "predecessor": f"{phase_index}.0",
                    "start": phase.get("start_date"),
                    "end": phase.get("end_date"),
                }
            )
        return tasks[:197]

    def _build_milestones(self, engine_milestones: List[Dict[str, Any]], phases: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        milestones = []
        source = engine_milestones or [
            {
                "milestone": f"{phase.get('phase')} Complete",
                "date": phase.get("end_date"),
                "phase": phase.get("phase"),
                "status": "Planned",
                "priority": "High",
            }
            for phase in phases
        ]
        for milestone in source[:38]:
            milestones.append(
                {
                    "milestone": milestone.get("milestone", "Milestone"),
                    "date": milestone.get("date"),
                    "phase": milestone.get("phase", ""),
                    "status": milestone.get("status", "Planned"),
                    "priority": milestone.get("priority", "High"),
                }
            )
        return milestones

    def _build_risks(self) -> List[Dict[str, Any]]:
        risks = []
        for risk in self.db_summary.get("risks", []):
            severity = (risk.get("severity") or "Medium").lower()
            impact = 5 if "high" in severity or "critical" in severity else 3 if "medium" in severity else 2
            risks.append(
                {
                    "description": risk.get("risk_description") or "Project risk",
                    "impact": impact,
                    "probability": 3,
                    "owner": "PM",
                    "mitigation": risk.get("mitigation") or "Track mitigation plan during weekly governance.",
                    "status": "Open" if risk.get("status", "Identified") != "Closed" else "Closed",
                }
            )

        if risks:
            return risks

        return [
            {
                "description": "Scope changes may affect delivery timeline",
                "impact": 4,
                "probability": 3,
                "owner": "PM",
                "mitigation": "Use Change_Log for approvals and impact assessment before implementation.",
                "status": "Open",
            },
            {
                "description": "Resource availability may constrain schedule",
                "impact": 3,
                "probability": 3,
                "owner": "PM",
                "mitigation": "Review Resource_Plan weekly and adjust allocation early.",
                "status": "Open",
            },
        ]

    def _build_assumptions(self) -> List[str]:
        return [
            "Core team is available for planned allocation.",
            "Client stakeholders are available for timely reviews and approvals.",
            "Required source systems, environments, and documentation are accessible.",
            "Scope baseline is approved before execution work begins.",
            "Budget baseline is approved and tracked in Budget_Tracker.",
        ]

    def _build_dependencies(self, engine_dependencies: List[Dict[str, Any]], phases: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        if engine_dependencies:
            return engine_dependencies
        return [
            {
                "task": phases[index + 1].get("phase"),
                "depends_on": phases[index].get("phase"),
                "dependency_type": "Finish to Start",
            }
            for index in range(max(0, len(phases) - 1))
        ]

    def _fallback_phases(self, start: datetime, duration_weeks: int) -> List[Dict[str, Any]]:
        names = ["Initiation", "Discovery", "Design", "Development", "Testing", "Deployment", "Closure"]
        weeks = max(1, duration_weeks // len(names))
        phases = []
        current = start
        for name in names:
            end = current + timedelta(weeks=weeks)
            phases.append(
                {
                    "phase": name,
                    "start_date": current.strftime("%Y-%m-%d"),
                    "end_date": end.strftime("%Y-%m-%d"),
                    "duration_weeks": weeks,
                    "status": "Planned",
                }
            )
            current = end
        return phases

    def _owner_for_phase(self, phase_name: str) -> str:
        lower = (phase_name or "").lower()
        if "requirement" in lower or "discovery" in lower:
            return "BA"
        if "design" in lower or "architecture" in lower or "model" in lower:
            return "Architect"
        if "development" in lower or "integration" in lower:
            return "Dev Lead"
        if "test" in lower or "validation" in lower or "uat" in lower:
            return "QA Lead"
        if "deploy" in lower or "migration" in lower:
            return "DevOps"
        return "PM"

    def _raci_value(self, activity: str, role: str, col_idx: int) -> str:
        lower = activity.lower()
        if role in ("PM", "Project Manager"):
            return "A"
        if role in ("BA", "Business Analyst") and ("requirement" in lower or "discovery" in lower):
            return "R"
        if role in ("Architect", "Tech Lead") and ("design" in lower or "architecture" in lower):
            return "R"
        if role in ("Developer", "Dev Lead") and ("development" in lower or "build" in lower or "integration" in lower):
            return "R"
        if role in ("QA Engineer", "QA Lead") and ("test" in lower or "uat" in lower or "validation" in lower):
            return "R"
        if role == "DevOps" and ("deploy" in lower or "go-live" in lower):
            return "R"
        if role == "Client PM":
            return "C"
        return "I" if col_idx % 3 == 0 else "C"

    def _midpoint_date(self, start: Any, end: Any) -> str:
        start_dt = self._parse_date(start)
        end_dt = self._parse_date(end)
        if not start_dt or not end_dt:
            return start
        return (start_dt + (end_dt - start_dt) / 2).strftime("%Y-%m-%d")

    def _rag_message(self, open_risks: int) -> str:
        if open_risks >= 5:
            return "RED - review RAID register"
        if open_risks:
            return "AMBER - review RAID register"
        return "GREEN - no major risks identified"

    def _resource_message(self, team: List[Dict[str, str]]) -> str:
        if len(team) >= 10:
            return "AMBER - review resource allocation"
        return "GREEN - baseline allocation ready"

    def _replace_list_validation(self, ws: Worksheet, range_ref: str, csv_options: str) -> None:
        from openpyxl.worksheet.datavalidation import DataValidation

        # Keep any unrelated validations, replace only same target ranges when possible.
        remaining = []
        for validation in ws.data_validations.dataValidation:
            if str(validation.sqref) != range_ref:
                remaining.append(validation)
        ws.data_validations.dataValidation = remaining

        validation = DataValidation(type="list", formula1=f'"{csv_options}"', allow_blank=True)
        ws.add_data_validation(validation)
        for part in range_ref.split():
            validation.add(part)

    def _copy_row_style(self, ws: Worksheet, source_row: int, target_row: int, start_col: int, end_col: int) -> None:
        for col in range(start_col, end_col + 1):
            source = ws.cell(source_row, col)
            target = ws.cell(target_row, col)
            if source.has_style:
                target._style = copy(source._style)
            if source.number_format:
                target.number_format = source.number_format
            if source.alignment:
                target.alignment = copy(source.alignment)
            if source.font:
                target.font = copy(source.font)
            if source.fill:
                target.fill = copy(source.fill)
            if source.border:
                target.border = copy(source.border)

    def _clear_values(self, ws: Worksheet, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row, col)
                if not type(cell).__name__ == "MergedCell":
                    cell.value = None

    def _format_dates(self, ws: Worksheet, cell_refs: List[str]) -> None:
        for ref in cell_refs:
            ws[ref].number_format = "yyyy-mm-dd"

    def _parse_date(self, value: Any) -> datetime:
        if isinstance(value, datetime):
            return value
        if not value:
            return None
        for date_format in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(str(value), date_format)
            except ValueError:
                continue
        return None

    def _safe_int(self, value: Any, default: int) -> int:
        try:
            return int(value)
        except (TypeError, ValueError):
            return default


class SampleFormatWorkbookFactory:
    """Factory for the approved sample-format workbook generator."""

    @staticmethod
    def create_sample_format_generator(project_info: Dict[str, Any], db_summary: Dict[str, Any]):
        return SampleFormatWorkbookGenerator(project_info, db_summary)
