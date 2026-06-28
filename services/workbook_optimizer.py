"""
PMO-Grade Workbook Optimizer for Project Aura.
Consolidates sheets, removes duplication, and creates a professional consulting deliverable.
"""

import logging
from datetime import datetime, timedelta
from typing import Dict, Any, List, Tuple
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from services.excel_formatter import ExcelFormatter
from services.project_plan_engine import ProjectPlanEngine

logger = logging.getLogger(__name__)

# Corporate consulting color scheme
COLORS = {
    'primary_dark': '1F4E78',      # Dark Navy
    'primary': '366092',            # Blue
    'header_light': 'D9EAF7',      # Light Blue
    'success': '70AD47',            # Green
    'warning': 'FFC000',            # Amber
    'danger': 'C5504F',             # Red
    'info': '4472C4',               # Info Blue
    'neutral': 'BFBFBF',            # Gray
    'white': 'FFFFFF',
    'black': '000000',
    'light_gray': 'F2F2F2',
    'row_alt': 'F8F8F8'
}

THIN_BORDER = Border(
    left=Side(style='thin', color='D3D3D3'),
    right=Side(style='thin', color='D3D3D3'),
    top=Side(style='thin', color='D3D3D3'),
    bottom=Side(style='thin', color='D3D3D3')
)


class PMOWorkbookOptimizer:
    """Generates PMO-grade, executive-ready workbooks with consolidated sheets."""

    @staticmethod
    def _format_title(ws, row: int, col: int, title: str, color: str = None):
        """Format a professional title."""
        cell = ws.cell(row=row, column=col)
        cell.value = title
        cell.font = Font(bold=True, size=16, color=COLORS['white'])
        cell.fill = PatternFill(start_color=color or COLORS['primary_dark'],
                              end_color=color or COLORS['primary_dark'], fill_type='solid')
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        return cell

    @staticmethod
    def _format_section_header(ws, row: int, col: int, header: str):
        """Format section headers."""
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True, size=12, color=COLORS['primary_dark'])
        cell.fill = PatternFill(start_color=COLORS['header_light'],
                              end_color=COLORS['header_light'], fill_type='solid')
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = THIN_BORDER
        return cell

    @staticmethod
    def _format_label(ws, row: int, col: int, label: str):
        """Format label cells."""
        cell = ws.cell(row=row, column=col)
        cell.value = label
        cell.font = Font(bold=True, size=11, color=COLORS['primary_dark'])
        cell.fill = PatternFill(start_color=COLORS['header_light'],
                              end_color=COLORS['header_light'], fill_type='solid')
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = THIN_BORDER
        return cell

    @staticmethod
    def _format_value(ws, row: int, col: int, value: Any):
        """Format value cells."""
        cell = ws.cell(row=row, column=col)
        cell.value = value
        cell.font = Font(size=11, color=COLORS['black'])
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
        return cell

    @staticmethod
    def _calculate_end_date(project_info: Dict) -> str:
        """Calculate end date from start date and duration."""
        try:
            start_date = project_info.get('start_date', '')
            duration = project_info.get('duration_weeks', 0)
            if start_date and duration:
                from datetime import datetime, timedelta
                start = datetime.strptime(str(start_date), '%Y-%m-%d')
                end = start + timedelta(weeks=int(duration))
                return end.strftime('%Y-%m-%d')
        except:
            pass
        return 'TBD'

    @staticmethod
    def _calculate_project_health(db_summary: Dict) -> Tuple[str, str]:
        """Calculate project health status and color."""
        risk_count = len(db_summary.get('risks', []))
        issue_count = len(db_summary.get('issues', []))
        dependencies = len(db_summary.get('dependencies', []))
        team_size = len(db_summary.get('team_members', []))
        complexity = risk_count + issue_count + dependencies

        if risk_count > 5 or issue_count > 3 or complexity > 15:
            return 'RED - High Risk', COLORS['danger']
        elif risk_count > 2 or issue_count > 1 or complexity > 8:
            return 'AMBER - Medium Risk', COLORS['warning']
        else:
            return 'GREEN - Low Risk', COLORS['success']

    @staticmethod
    def _calculate_confidence_score(project_info: Dict, db_summary: Dict) -> int:
        """Calculate delivery confidence score (0-100)."""
        score = 100

        # Reduce score for risks
        score -= min(len(db_summary.get('risks', [])) * 5, 30)

        # Reduce score for dependencies
        score -= min(len(db_summary.get('dependencies', [])) * 2, 15)

        # Reduce score for large teams
        team_size = project_info.get('team_size', 5)
        if team_size > 10:
            score -= (team_size - 10) * 2

        return max(score, 20)

    @staticmethod
    def create_home_page(workbook, project_info: Dict, db_summary: Dict):
        """Create professional Home/Navigation page."""
        ws = workbook.create_sheet('00_Home', 0)
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 40

        # Title
        ws.merge_cells('A1:B1')
        cell = ws['A1']
        cell.value = 'PROJECT AURA'
        cell.font = Font(bold=True, size=20, color=COLORS['white'])
        cell.fill = PatternFill(start_color=COLORS['primary_dark'],
                              end_color=COLORS['primary_dark'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 35

        # Subtitle
        ws.merge_cells('A2:B2')
        cell = ws['A2']
        cell.value = 'AI-Generated Project Management Workbook'
        cell.font = Font(italic=True, size=12, color='666666')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 20

        # Project Summary
        row = 4
        PMOWorkbookOptimizer._format_section_header(ws, row, 1, 'PROJECT SUMMARY')
        ws.merge_cells(f'A{row}:B{row}')

        row += 1
        PMOWorkbookOptimizer._format_label(ws, row, 1, 'Project Name')
        PMOWorkbookOptimizer._format_value(ws, row, 2, project_info.get('project_name', 'N/A'))

        row += 1
        PMOWorkbookOptimizer._format_label(ws, row, 1, 'Client Name')
        PMOWorkbookOptimizer._format_value(ws, row, 2, project_info.get('client_name', 'N/A'))

        row += 1
        PMOWorkbookOptimizer._format_label(ws, row, 1, 'Project Type')
        PMOWorkbookOptimizer._format_value(ws, row, 2, project_info.get('project_type', 'N/A'))

        row += 1
        PMOWorkbookOptimizer._format_label(ws, row, 1, 'Start Date')
        PMOWorkbookOptimizer._format_value(ws, row, 2, project_info.get('start_date', 'TBD'))

        row += 1
        PMOWorkbookOptimizer._format_label(ws, row, 1, 'End Date')
        PMOWorkbookOptimizer._format_value(ws, row, 2,
                                          PMOWorkbookOptimizer._calculate_end_date(project_info))

        row += 1
        PMOWorkbookOptimizer._format_label(ws, row, 1, 'Duration')
        PMOWorkbookOptimizer._format_value(ws, row, 2,
                                          f"{project_info.get('duration_weeks', 0)} weeks")

        row += 1
        PMOWorkbookOptimizer._format_label(ws, row, 1, 'Estimated Team Size')
        PMOWorkbookOptimizer._format_value(ws, row, 2, project_info.get('team_size', 0))

        row += 1
        PMOWorkbookOptimizer._format_label(ws, row, 1, 'Delivery Model')
        PMOWorkbookOptimizer._format_value(ws, row, 2, project_info.get('delivery_model', 'N/A'))

        row += 1
        PMOWorkbookOptimizer._format_label(ws, row, 1, 'Generated Date')
        PMOWorkbookOptimizer._format_value(ws, row, 2, datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

        # Navigation section
        row += 2
        PMOWorkbookOptimizer._format_section_header(ws, row, 1, 'NAVIGATION')
        ws.merge_cells(f'A{row}:B{row}')

        navigation_items = [
            ('01_Executive_Dashboard', 'Executive KPIs and project health'),
            ('02_AI_Project_Summary', 'AI-generated insights and recommendations'),
            ('03_Project_Details', 'Project charter and detailed information'),
            ('04_Project_Roadmap', 'Phase-level timeline and milestones'),
            ('05_Detailed_Project_Plan', 'Task-level project plan'),
            ('06_Gantt_Chart', 'Visual timeline and critical path'),
            ('07_Milestone_Tracker', 'Key milestones and deliverables'),
            ('08_Resource_Plan', 'Resource allocation and capacity'),
            ('09_RAID_Register', 'Risks, assumptions, issues, dependencies'),
            ('10_RACI_Matrix', 'Responsibility assignments'),
            ('11_Weekly_Status', 'Progress tracking and status')
        ]

        row += 1
        for sheet_name, description in navigation_items:
            cell = ws.cell(row=row, column=1)
            cell.value = sheet_name
            cell.font = Font(underline='single', color='0563C1', size=11)
            cell.alignment = Alignment(horizontal='left', vertical='center')

            cell = ws.cell(row=row, column=2)
            cell.value = description
            cell.font = Font(size=11, color='666666')
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

            row += 1

        # Footer
        row += 1
        ws.merge_cells(f'A{row}:B{row}')
        cell = ws[f'A{row}']
        cell.value = f"Generated by Project Aura | {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        cell.font = Font(italic=True, size=9, color='999999')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    @staticmethod
    def create_executive_dashboard(workbook, project_info: Dict, db_summary: Dict):
        """Create professional executive dashboard with KPI cards."""
        ws = workbook.create_sheet('01_Executive_Dashboard', 1)

        # Title
        ws.merge_cells('A1:H1')
        cell = ws['A1']
        cell.value = 'EXECUTIVE PROJECT DASHBOARD'
        cell.font = Font(bold=True, size=16, color=COLORS['white'])
        cell.fill = PatternFill(start_color=COLORS['primary_dark'],
                              end_color=COLORS['primary_dark'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        # Timestamp
        ws.merge_cells('A2:H2')
        cell = ws['A2']
        cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        cell.font = Font(size=10, italic=True, color='666666')
        cell.alignment = Alignment(horizontal='right')

        # KPI Cards
        row = 4
        kpis = [
            ('Project Duration', f"{project_info.get('duration_weeks', 0)} weeks", COLORS['info']),
            ('Team Size', f"{project_info.get('team_size', 0)} people", COLORS['info']),
            ('Risk Count', len(db_summary.get('risks', [])), COLORS['danger']),
            ('Deliverables', len(db_summary.get('deliverables', [])), COLORS['success']),
        ]

        col = 1
        for label, value, color in kpis:
            # Card header
            ws.merge_cells(f'{get_column_letter(col)}{row}:{get_column_letter(col)}{row}')
            cell = ws.cell(row=row, column=col)
            cell.value = label
            cell.font = Font(bold=True, size=11, color=COLORS['white'])
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # Card value
            cell = ws.cell(row=row + 1, column=col)
            cell.value = value
            cell.font = Font(bold=True, size=14)
            cell.fill = PatternFill(start_color=COLORS['light_gray'],
                                  end_color=COLORS['light_gray'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

            col += 2

        # Health Indicator
        row += 3
        ws.merge_cells(f'A{row}:H{row}')
        cell = ws[f'A{row}']
        cell.value = 'PROJECT HEALTH'
        cell.font = Font(bold=True, size=12, color=COLORS['primary_dark'])
        cell.fill = PatternFill(start_color=COLORS['header_light'],
                              end_color=COLORS['header_light'], fill_type='solid')
        cell.alignment = Alignment(horizontal='left', vertical='center')

        row += 1
        health_status, health_color = PMOWorkbookOptimizer._calculate_project_health(db_summary)
        ws.merge_cells(f'A{row}:H{row}')
        cell = ws[f'A{row}']
        cell.value = health_status
        cell.font = Font(bold=True, size=14, color=COLORS['white'])
        cell.fill = PatternFill(start_color=health_color, end_color=health_color, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 25

        # Summary metrics
        row += 2
        ws.merge_cells(f'A{row}:H{row}')
        cell = ws[f'A{row}']
        cell.value = 'KEY METRICS'
        cell.font = Font(bold=True, size=12, color=COLORS['primary_dark'])
        cell.fill = PatternFill(start_color=COLORS['header_light'],
                              end_color=COLORS['header_light'], fill_type='solid')

        metrics = [
            ('Total Team Members', len(db_summary.get('team_members', []))),
            ('Total Risks', len(db_summary.get('risks', []))),
            ('Total Dependencies', len(db_summary.get('dependencies', []))),
            ('Confidence Score', f"{PMOWorkbookOptimizer._calculate_confidence_score(project_info, db_summary)}%"),
        ]

        row += 1
        col_widths = {
            'A': 30, 'B': 20, 'C': 30, 'D': 20, 'E': 30, 'F': 20, 'G': 30, 'H': 20
        }
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        for i, (label, value) in enumerate(metrics):
            if i % 2 == 0:
                col = 1
                if i > 0:
                    row += 1
            else:
                col = 5

            PMOWorkbookOptimizer._format_label(ws, row, col, label)
            PMOWorkbookOptimizer._format_value(ws, row, col + 1, value)

        row += 2

        # Top Risks
        ws.merge_cells(f'A{row}:H{row}')
        cell = ws[f'A{row}']
        cell.value = 'TOP RISKS'
        cell.font = Font(bold=True, size=12, color=COLORS['primary_dark'])
        cell.fill = PatternFill(start_color=COLORS['header_light'],
                              end_color=COLORS['header_light'], fill_type='solid')

        row += 1
        headers = ['Risk', 'Severity', 'Mitigation']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True, size=11, color=COLORS['white'])
            cell.fill = PatternFill(start_color=COLORS['primary'],
                                  end_color=COLORS['primary'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        row += 1
        for risk in db_summary.get('risks', [])[:5]:
            PMOWorkbookOptimizer._format_value(ws, row, 1, risk.get('risk_description', ''))
            severity = risk.get('severity', 'Low')
            color = COLORS['danger'] if severity == 'High' else COLORS['warning']
            cell = ws.cell(row=row, column=2)
            cell.value = severity
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            cell.font = Font(color=COLORS['white'], bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            PMOWorkbookOptimizer._format_value(ws, row, 3, risk.get('mitigation', ''))
            row += 1

        ExcelFormatter.freeze_panes(ws, 3)

    @staticmethod
    def create_ai_project_summary(workbook, project_info: Dict, db_summary: Dict):
        """Create AI-powered executive briefing with insights and recommendations."""
        ws = workbook.create_sheet('02_AI_Project_Summary', 2)
        ws.column_dimensions['A'].width = 100

        # Title
        ws.merge_cells('A1:A1')
        cell = ws['A1']
        cell.value = 'AI PROJECT SUMMARY & RECOMMENDATIONS'
        cell.font = Font(bold=True, size=16, color=COLORS['white'])
        cell.fill = PatternFill(start_color=COLORS['primary_dark'],
                              end_color=COLORS['primary_dark'], fill_type='solid')
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.row_dimensions[1].height = 25

        row = 3
        duration = project_info.get('duration_weeks', 0)
        team_size = project_info.get('team_size', 0)
        project_type = project_info.get('project_type', 'Project')
        risk_count = len(db_summary.get('risks', []))

        # Executive narrative
        ws.merge_cells(f'A{row}:A{row}')
        cell = ws[f'A{row}']
        cell.value = 'EXECUTIVE NARRATIVE'
        cell.font = Font(bold=True, size=12, color=COLORS['primary_dark'])
        cell.fill = PatternFill(start_color=COLORS['header_light'],
                              end_color=COLORS['header_light'], fill_type='solid')

        row += 1
        narrative = f"""Based on the uploaded Statement of Work, this initiative is classified as a {project_type} project requiring approximately {team_size} resources over {duration} weeks with an estimated effort of {duration * 40 * team_size} hours.

Key characteristics:
• Project Type: {project_type}
• Estimated Duration: {duration} weeks ({duration/4:.1f} months)
• Estimated Team Size: {team_size} people
• Estimated Effort: {duration * 40 * team_size} hours
• Identified Risks: {risk_count}
• Project Complexity: {'High' if risk_count > 5 else 'Medium' if risk_count > 2 else 'Low'}"""

        ws.merge_cells(f'A{row}:A{row + 5}')
        cell = ws[f'A{row}']
        cell.value = narrative
        cell.font = Font(size=11, color='000000')
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        row += 6

        # Recommendations
        ws.merge_cells(f'A{row}:A{row}')
        cell = ws[f'A{row}']
        cell.value = 'RECOMMENDED APPROACH'
        cell.font = Font(bold=True, size=12, color=COLORS['primary_dark'])
        cell.fill = PatternFill(start_color=COLORS['header_light'],
                              end_color=COLORS['header_light'], fill_type='solid')

        row += 1

        methodology = 'Agile with Scrum' if duration < 24 else 'Agile with Kanban' if risk_count < 3 else 'Hybrid (Agile + Waterfall)'
        governance = 'Lightweight PMO' if team_size <= 5 else 'Standard PMO' if team_size <= 10 else 'Enterprise PMO'

        recommendations = f"""• Recommended Delivery Model: {methodology}
• Recommended Governance: {governance}
• Recommended Team Structure: {team_size} FTE across Development, QA, and Operations
• Critical Success Factors: Resource availability, stakeholder alignment, scope management
• Top Priority Risks to Mitigate: {', '.join([r.get('risk_description', 'Risk') for r in db_summary.get('risks', [])[:2]])}
• Estimated Project Budget: Based on effort and standard rates
• Recommended Cadence: Weekly status meetings, bi-weekly reviews, monthly steering committee"""

        ws.merge_cells(f'A{row}:A{row + 5}')
        cell = ws[f'A{row}']
        cell.value = recommendations
        cell.font = Font(size=11, color='000000')
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        row += 6

        # Metrics
        row += 1
        metrics = [
            ('Confidence Score', f"{PMOWorkbookOptimizer._calculate_confidence_score(project_info, db_summary)}%"),
            ('Risk Exposure', 'High' if risk_count > 5 else 'Medium' if risk_count > 2 else 'Low'),
            ('Complexity Rating', 'High' if risk_count > 5 or team_size > 10 else 'Medium' if risk_count > 2 or team_size > 5 else 'Low'),
        ]

        ws.merge_cells(f'A{row}:A{row}')
        cell = ws[f'A{row}']
        cell.value = 'DELIVERY METRICS'
        cell.font = Font(bold=True, size=12, color=COLORS['primary_dark'])
        cell.fill = PatternFill(start_color=COLORS['header_light'],
                              end_color=COLORS['header_light'], fill_type='solid')

        row += 1
        for label, value in metrics:
            ws.merge_cells(f'A{row}:A{row}')
            cell = ws[f'A{row}']
            cell.value = f'{label}: {value}'
            cell.font = Font(size=11, bold=True)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            row += 1

    @staticmethod
    def create_project_details(workbook, project_info: Dict, db_summary: Dict):
        """Create consolidated Project Details sheet (merged Charter + Details)."""
        ws = workbook.create_sheet('03_Project_Details', 3)
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 70

        row = PMOWorkbookOptimizer._add_section(ws, row=1, title='PROJECT OVERVIEW')

        details = [
            ('Project Name', project_info.get('project_name', 'N/A')),
            ('Client Name', project_info.get('client_name', 'N/A')),
            ('Project Type', project_info.get('project_type', 'N/A')),
            ('Delivery Model', project_info.get('delivery_model', 'N/A')),
            ('Start Date', project_info.get('start_date', 'TBD')),
            ('End Date', PMOWorkbookOptimizer._calculate_end_date(project_info)),
            ('Duration', f"{project_info.get('duration_weeks', 0)} weeks"),
            ('Team Size', project_info.get('team_size', 0)),
        ]

        for label, value in details:
            PMOWorkbookOptimizer._format_label(ws, row, 1, label)
            PMOWorkbookOptimizer._format_value(ws, row, 2, value)
            row += 1

        row += 1
        row = PMOWorkbookOptimizer._add_section(ws, row=row, title='PROJECT SCOPE')

        PMOWorkbookOptimizer._format_label(ws, row, 1, 'Scope Statement')
        row += 1
        ws.merge_cells(f'A{row}:B{row + 2}')
        cell = ws[f'A{row}']
        cell.value = project_info.get('scope', 'Not provided')
        cell.font = Font(size=11)
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.row_dimensions[row].height = 45
        row += 3

        row += 1
        row = PMOWorkbookOptimizer._add_section(ws, row=row, title='ASSUMPTIONS')

        if db_summary.get('assumptions'):
            for i, assumption in enumerate(db_summary.get('assumptions', []), 1):
                PMOWorkbookOptimizer._format_label(ws, row, 1, f'Assumption {i}')
                PMOWorkbookOptimizer._format_value(ws, row, 2, assumption)
                row += 1
        else:
            PMOWorkbookOptimizer._format_value(ws, row, 1, 'No assumptions defined')
            row += 1

        row += 1
        row = PMOWorkbookOptimizer._add_section(ws, row=row, title='CONSTRAINTS')

        constraints = [
            'Budget constraints as per contract',
            'Schedule constraints based on delivery model',
            'Resource availability and allocation',
            'Third-party dependencies',
            'Technology and platform limitations',
        ]

        for constraint in constraints:
            PMOWorkbookOptimizer._format_value(ws, row, 1, '• ' + constraint)
            ws.merge_cells(f'A{row}:B{row}')
            row += 1

        ExcelFormatter.freeze_panes(ws, 3)

    @staticmethod
    def _add_section(ws, row: int, title: str) -> int:
        """Add a section header and return next row."""
        ws.merge_cells(f'A{row}:B{row}')
        cell = ws[f'A{row}']
        cell.value = title
        cell.font = Font(bold=True, size=12, color=COLORS['white'])
        cell.fill = PatternFill(start_color=COLORS['primary'],
                              end_color=COLORS['primary'], fill_type='solid')
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 20
        return row + 1

    @staticmethod
    def create_project_roadmap(workbook, project_info: Dict):
        """Create phase-level roadmap with timeline visualization."""
        ws = workbook.create_sheet('04_Project_Roadmap', 4)

        # Title
        ws.merge_cells('A1:H1')
        cell = ws['A1']
        cell.value = 'PROJECT ROADMAP'
        cell.font = Font(bold=True, size=14, color=COLORS['white'])
        cell.fill = PatternFill(start_color=COLORS['primary_dark'],
                              end_color=COLORS['primary_dark'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Headers
        headers = ['Phase', 'Start Date', 'End Date', 'Duration', 'Status', 'Owner', 'Key Deliverables', 'Notes']
        row = 3
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True, size=11, color=COLORS['white'])
            cell.fill = PatternFill(start_color=COLORS['primary'],
                                  end_color=COLORS['primary'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Phases
        phases = ['Initiation', 'Discovery', 'Design', 'Development', 'Testing', 'UAT', 'Deployment', 'Hypercare']
        start_date = project_info.get('start_date', datetime.now().strftime('%Y-%m-%d'))
        duration_weeks = project_info.get('duration_weeks', 16)
        weeks_per_phase = max(1, duration_weeks // len(phases))

        try:
            current_start = datetime.strptime(str(start_date), '%Y-%m-%d')
        except:
            current_start = datetime.now()

        row = 4
        for phase in phases:
            phase_end = current_start + timedelta(weeks=weeks_per_phase)
            PMOWorkbookOptimizer._format_value(ws, row, 1, phase)
            PMOWorkbookOptimizer._format_value(ws, row, 2, current_start.strftime('%Y-%m-%d'))
            PMOWorkbookOptimizer._format_value(ws, row, 3, phase_end.strftime('%Y-%m-%d'))
            PMOWorkbookOptimizer._format_value(ws, row, 4, f'{weeks_per_phase}w')

            status_cell = ws.cell(row=row, column=5)
            status_cell.value = 'Planned'
            status_cell.fill = PatternFill(start_color=COLORS['neutral'],
                                         end_color=COLORS['neutral'], fill_type='solid')
            status_cell.alignment = Alignment(horizontal='center', vertical='center')

            PMOWorkbookOptimizer._format_value(ws, row, 6, 'TBD')
            PMOWorkbookOptimizer._format_value(ws, row, 7, 'TBD')
            PMOWorkbookOptimizer._format_value(ws, row, 8, '')

            current_start = phase_end
            row += 1

        col_widths = {'A': 15, 'B': 15, 'C': 15, 'D': 12, 'E': 12, 'F': 15, 'G': 20, 'H': 20}
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        ExcelFormatter.freeze_panes(ws, 4)
        ExcelFormatter.add_filter(ws, 'A3', f'{get_column_letter(len(headers))}100')

    @staticmethod
    def create_detailed_project_plan(workbook, project_info: Dict, db_summary: Dict):
        """Create master project planning sheet with comprehensive task details."""
        ws = workbook.create_sheet('05_Detailed_Project_Plan', 5)

        # Title
        ws.merge_cells('A1:M1')
        cell = ws['A1']
        cell.value = 'DETAILED PROJECT PLAN'
        cell.font = Font(bold=True, size=14, color=COLORS['white'])
        cell.fill = PatternFill(start_color=COLORS['primary_dark'],
                              end_color=COLORS['primary_dark'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Headers
        headers = ['WBS', 'Phase', 'Task', 'Deliverable', 'Owner', 'Start Date', 'End Date',
                  'Duration', 'Dependency', 'Priority', 'Status', 'Completion %', 'Comments']
        row = 3
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True, size=10, color=COLORS['white'])
            cell.fill = PatternFill(start_color=COLORS['primary'],
                                  end_color=COLORS['primary'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.column_dimensions[get_column_letter(col)].width = 12

        row = 4
        tasks = [
            ('1.0', 'Initiation', 'Project kickoff', 'Kickoff memo', 'PM', 0, 1, 'High'),
            ('2.0', 'Discovery', 'Requirements gathering', 'Requirements doc', 'BA', 1, 3, 'High'),
            ('3.0', 'Design', 'Solution design', 'Design spec', 'Architect', 3, 5, 'High'),
            ('4.0', 'Development', 'Development', 'Code release', 'Dev Lead', 5, 10, 'High'),
            ('5.0', 'Testing', 'QA testing', 'Test report', 'QA Lead', 10, 12, 'Medium'),
            ('6.0', 'UAT', 'User acceptance', 'UAT sign-off', 'Client PM', 12, 13, 'High'),
            ('7.0', 'Deployment', 'Production deployment', 'Go-live report', 'DevOps', 13, 14, 'High'),
            ('8.0', 'Hypercare', 'Support phase', 'Hypercare report', 'PM', 14, 16, 'Medium'),
        ]

        for wbs, phase, task, deliverable, owner, start_week, end_week, priority in tasks:
            try:
                start_date = datetime.strptime(str(project_info.get('start_date', '')), '%Y-%m-%d') + timedelta(weeks=start_week)
                end_date = datetime.strptime(str(project_info.get('start_date', '')), '%Y-%m-%d') + timedelta(weeks=end_week)
            except:
                start_date = 'TBD'
                end_date = 'TBD'

            PMOWorkbookOptimizer._format_value(ws, row, 1, wbs)
            PMOWorkbookOptimizer._format_value(ws, row, 2, phase)
            PMOWorkbookOptimizer._format_value(ws, row, 3, task)
            PMOWorkbookOptimizer._format_value(ws, row, 4, deliverable)
            PMOWorkbookOptimizer._format_value(ws, row, 5, owner)
            PMOWorkbookOptimizer._format_value(ws, row, 6, start_date if isinstance(start_date, str) else start_date.strftime('%Y-%m-%d'))
            PMOWorkbookOptimizer._format_value(ws, row, 7, end_date if isinstance(end_date, str) else end_date.strftime('%Y-%m-%d'))
            PMOWorkbookOptimizer._format_value(ws, row, 8, f'{end_week - start_week}w')
            PMOWorkbookOptimizer._format_value(ws, row, 9, 'None')

            priority_cell = ws.cell(row=row, column=10)
            priority_cell.value = priority
            priority_cell.fill = PatternFill(start_color=COLORS['warning'],
                                           end_color=COLORS['warning'], fill_type='solid')
            priority_cell.alignment = Alignment(horizontal='center', vertical='center')

            status_cell = ws.cell(row=row, column=11)
            status_cell.value = 'Planned'
            status_cell.fill = PatternFill(start_color=COLORS['neutral'],
                                         end_color=COLORS['neutral'], fill_type='solid')
            status_cell.alignment = Alignment(horizontal='center', vertical='center')

            PMOWorkbookOptimizer._format_value(ws, row, 12, '0%')
            PMOWorkbookOptimizer._format_value(ws, row, 13, '')

            row += 1

        ExcelFormatter.freeze_panes(ws, 4)
        ExcelFormatter.add_filter(ws, 'A3', 'M100')

    @staticmethod
    def create_gantt_chart(workbook, project_info: Dict):
        """Create professional Gantt chart visualization."""
        ws = workbook.create_sheet('06_Gantt_Chart', 6)

        # Title
        ws.merge_cells('A1:AA1')
        cell = ws['A1']
        cell.value = 'PROJECT GANTT CHART'
        cell.font = Font(bold=True, size=14, color=COLORS['white'])
        cell.fill = PatternFill(start_color=COLORS['primary_dark'],
                              end_color=COLORS['primary_dark'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Task names column
        ws.column_dimensions['A'].width = 20

        tasks = ['Initiation', 'Discovery', 'Design', 'Development', 'Testing', 'UAT', 'Deployment', 'Hypercare']

        row = 3
        cell = ws.cell(row=row, column=1)
        cell.value = 'Task'
        cell.font = Font(bold=True, size=10, color=COLORS['white'])
        cell.fill = PatternFill(start_color=COLORS['primary'],
                              end_color=COLORS['primary'], fill_type='solid')

        # Month headers
        start_date = datetime.strptime(str(project_info.get('start_date', '2025-01-01')), '%Y-%m-%d')
        duration_weeks = project_info.get('duration_weeks', 16)
        months = max(1, (duration_weeks // 4) + 1)

        for month_idx in range(months):
            month_date = start_date + timedelta(weeks=month_idx * 4)
            col = month_idx + 2
            cell = ws.cell(row=row, column=col)
            cell.value = month_date.strftime('%b %Y')
            cell.font = Font(bold=True, size=10, color=COLORS['white'])
            cell.fill = PatternFill(start_color=COLORS['primary'],
                                  end_color=COLORS['primary'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions[get_column_letter(col)].width = 12

        # Task bars
        row = 4
        duration_per_task = duration_weeks // len(tasks)

        for task in tasks:
            cell = ws.cell(row=row, column=1)
            cell.value = task
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color=COLORS['header_light'],
                                  end_color=COLORS['header_light'], fill_type='solid')
            cell.alignment = Alignment(horizontal='left', vertical='center')

            # Draw task bar
            for month_idx in range(months):
                col = month_idx + 2
                cell = ws.cell(row=row, column=col)
                cell.fill = PatternFill(start_color=COLORS['info'],
                                      end_color=COLORS['info'], fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.value = '████'

            row += 1

        ExcelFormatter.freeze_panes(ws, 3)

    @staticmethod
    def create_milestone_tracker(workbook, project_info: Dict, db_summary: Dict):
        """Create milestone tracking sheet with key deliverables."""
        ws = workbook.create_sheet('07_Milestone_Tracker', 7)

        # Title
        ws.merge_cells('A1:H1')
        cell = ws['A1']
        cell.value = 'MILESTONE TRACKER'
        cell.font = Font(bold=True, size=14, color=COLORS['white'])
        cell.fill = PatternFill(start_color=COLORS['primary_dark'],
                              end_color=COLORS['primary_dark'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Headers
        headers = ['Milestone ID', 'Description', 'Owner', 'Planned Date', 'Actual Date', 'Status', 'Priority', 'Remarks']
        row = 3
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True, size=11, color=COLORS['white'])
            cell.fill = PatternFill(start_color=COLORS['primary'],
                                  end_color=COLORS['primary'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Sample milestones
        milestones = [
            ('M1', 'Project Kickoff', 'PM', 0),
            ('M2', 'Requirements Complete', 'BA', 2),
            ('M3', 'Design Review', 'Architect', 4),
            ('M4', 'Development Complete', 'Dev Lead', 9),
            ('M5', 'Testing Complete', 'QA Lead', 11),
            ('M6', 'UAT Sign-off', 'Client PM', 12),
            ('M7', 'Go-Live', 'DevOps', 13),
            ('M8', 'Hypercare Complete', 'PM', 15),
        ]

        row = 4
        for milestone_id, description, owner, week in milestones:
            try:
                milestone_date = datetime.strptime(str(project_info.get('start_date', '')), '%Y-%m-%d') + timedelta(weeks=week)
                milestone_date_str = milestone_date.strftime('%Y-%m-%d')
            except:
                milestone_date_str = 'TBD'

            PMOWorkbookOptimizer._format_value(ws, row, 1, milestone_id)
            PMOWorkbookOptimizer._format_value(ws, row, 2, description)
            PMOWorkbookOptimizer._format_value(ws, row, 3, owner)
            PMOWorkbookOptimizer._format_value(ws, row, 4, milestone_date_str)
            PMOWorkbookOptimizer._format_value(ws, row, 5, '')

            status_cell = ws.cell(row=row, column=6)
            status_cell.value = 'Planned'
            status_cell.fill = PatternFill(start_color=COLORS['neutral'],
                                         end_color=COLORS['neutral'], fill_type='solid')
            status_cell.alignment = Alignment(horizontal='center', vertical='center')

            priority_cell = ws.cell(row=row, column=7)
            priority_cell.value = 'High'
            priority_cell.fill = PatternFill(start_color=COLORS['warning'],
                                           end_color=COLORS['warning'], fill_type='solid')
            priority_cell.alignment = Alignment(horizontal='center', vertical='center')

            PMOWorkbookOptimizer._format_value(ws, row, 8, '')
            row += 1

        col_widths = {'A': 12, 'B': 25, 'C': 15, 'D': 15, 'E': 15, 'F': 12, 'G': 12, 'H': 20}
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        ExcelFormatter.freeze_panes(ws, 4)
        ExcelFormatter.add_filter(ws, 'A3', 'H100')

    @staticmethod
    def create_resource_plan(workbook, project_info: Dict, db_summary: Dict):
        """Create PMO-style resource planning sheet."""
        ws = workbook.create_sheet('08_Resource_Plan', 8)

        # Title
        ws.merge_cells('A1:H1')
        cell = ws['A1']
        cell.value = 'RESOURCE PLAN'
        cell.font = Font(bold=True, size=14, color=COLORS['white'])
        cell.fill = PatternFill(start_color=COLORS['primary_dark'],
                              end_color=COLORS['primary_dark'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Headers
        headers = ['Resource', 'Role', 'Department', 'Allocation %', 'Start Date', 'End Date', 'Effort Hours', 'Utilization %']
        row = 3
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True, size=11, color=COLORS['white'])
            cell.fill = PatternFill(start_color=COLORS['primary'],
                                  end_color=COLORS['primary'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Sample resources
        row = 4
        for team_member in db_summary.get('team_members', []):
            role = team_member.get('role', 'Staff')
            count = team_member.get('count', 1)

            for i in range(count):
                duration = project_info.get('duration_weeks', 16)
                effort_hours = duration * 40

                PMOWorkbookOptimizer._format_value(ws, row, 1, f'{role}_{i+1}')
                PMOWorkbookOptimizer._format_value(ws, row, 2, role)
                PMOWorkbookOptimizer._format_value(ws, row, 3, 'Engineering')

                allocation_cell = ws.cell(row=row, column=4)
                allocation_cell.value = '100%'
                allocation_cell.alignment = Alignment(horizontal='center', vertical='center')

                PMOWorkbookOptimizer._format_value(ws, row, 5, project_info.get('start_date', 'TBD'))
                PMOWorkbookOptimizer._format_value(ws, row, 6, PMOWorkbookOptimizer._calculate_end_date(project_info))

                effort_cell = ws.cell(row=row, column=7)
                effort_cell.value = effort_hours
                effort_cell.alignment = Alignment(horizontal='right', vertical='center')

                util_cell = ws.cell(row=row, column=8)
                util_cell.value = '100%'
                util_cell.alignment = Alignment(horizontal='center', vertical='center')

                row += 1

        col_widths = {'A': 15, 'B': 15, 'C': 15, 'D': 14, 'E': 14, 'F': 14, 'G': 14, 'H': 14}
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        ExcelFormatter.freeze_panes(ws, 4)
        ExcelFormatter.add_filter(ws, 'A3', 'H100')

    @staticmethod
    def create_raid_register(workbook, project_info: Dict, db_summary: Dict):
        """Create unified RAID register (Risks, Assumptions, Issues, Dependencies)."""
        ws = workbook.create_sheet('09_RAID_Register', 9)

        # Title
        ws.merge_cells('A1:H1')
        cell = ws['A1']
        cell.value = 'RAID REGISTER (Risks, Assumptions, Issues, Dependencies)'
        cell.font = Font(bold=True, size=14, color=COLORS['white'])
        cell.fill = PatternFill(start_color=COLORS['primary_dark'],
                              end_color=COLORS['primary_dark'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Summary statistics
        row = 3
        stats = [
            ('Total Risks', len(db_summary.get('risks', []))),
            ('Total Assumptions', len(db_summary.get('assumptions', []))),
            ('Total Issues', 0),
            ('Total Dependencies', len(db_summary.get('dependencies', []))),
        ]

        for label, count in stats:
            PMOWorkbookOptimizer._format_label(ws, row, 1, label)
            PMOWorkbookOptimizer._format_value(ws, row, 2, count)
            row += 1

        # RAID Details
        row += 1
        headers = ['ID', 'Type', 'Description', 'Impact', 'Probability', 'Owner', 'Mitigation/Action', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True, size=11, color=COLORS['white'])
            cell.fill = PatternFill(start_color=COLORS['primary'],
                                  end_color=COLORS['primary'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        row += 1
        raid_id = 1

        # Add risks
        for risk in db_summary.get('risks', []):
            PMOWorkbookOptimizer._format_value(ws, row, 1, f'R{raid_id}')

            type_cell = ws.cell(row=row, column=2)
            type_cell.value = 'Risk'
            type_cell.fill = PatternFill(start_color=COLORS['danger'],
                                       end_color=COLORS['danger'], fill_type='solid')
            type_cell.font = Font(color=COLORS['white'], bold=True)
            type_cell.alignment = Alignment(horizontal='center', vertical='center')

            PMOWorkbookOptimizer._format_value(ws, row, 3, risk.get('risk_description', ''))
            PMOWorkbookOptimizer._format_value(ws, row, 4, 'High')
            PMOWorkbookOptimizer._format_value(ws, row, 5, risk.get('severity', 'Medium'))
            PMOWorkbookOptimizer._format_value(ws, row, 6, 'TBD')
            PMOWorkbookOptimizer._format_value(ws, row, 7, risk.get('mitigation', ''))

            status_cell = ws.cell(row=row, column=8)
            status_cell.value = 'Open'
            status_cell.fill = PatternFill(start_color=COLORS['warning'],
                                         end_color=COLORS['warning'], fill_type='solid')
            status_cell.alignment = Alignment(horizontal='center', vertical='center')

            raid_id += 1
            row += 1

        # Add assumptions
        for assumption in db_summary.get('assumptions', []):
            PMOWorkbookOptimizer._format_value(ws, row, 1, f'A{raid_id}')

            type_cell = ws.cell(row=row, column=2)
            type_cell.value = 'Assumption'
            type_cell.fill = PatternFill(start_color=COLORS['info'],
                                       end_color=COLORS['info'], fill_type='solid')
            type_cell.font = Font(color=COLORS['white'], bold=True)
            type_cell.alignment = Alignment(horizontal='center', vertical='center')

            PMOWorkbookOptimizer._format_value(ws, row, 3, assumption)
            PMOWorkbookOptimizer._format_value(ws, row, 4, 'Medium')
            PMOWorkbookOptimizer._format_value(ws, row, 5, 'Medium')
            PMOWorkbookOptimizer._format_value(ws, row, 6, 'TBD')
            PMOWorkbookOptimizer._format_value(ws, row, 7, 'Validate')

            status_cell = ws.cell(row=row, column=8)
            status_cell.value = 'Active'
            status_cell.fill = PatternFill(start_color=COLORS['success'],
                                         end_color=COLORS['success'], fill_type='solid')
            status_cell.alignment = Alignment(horizontal='center', vertical='center')

            raid_id += 1
            row += 1

        col_widths = {'A': 8, 'B': 12, 'C': 30, 'D': 10, 'E': 12, 'F': 12, 'G': 20, 'H': 12}
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        ExcelFormatter.freeze_panes(ws, 9)
        ExcelFormatter.add_filter(ws, 'A9', 'H200')

    @staticmethod
    def create_raci_matrix(workbook, project_info: Dict, db_summary: Dict):
        """Create RACI matrix for responsibility assignments."""
        ws = workbook.create_sheet('10_RACI_Matrix', 10)

        # Title
        ws.merge_cells('A1:H1')
        cell = ws['A1']
        cell.value = 'RACI MATRIX (Responsible, Accountable, Consulted, Informed)'
        cell.font = Font(bold=True, size=14, color=COLORS['white'])
        cell.fill = PatternFill(start_color=COLORS['primary_dark'],
                              end_color=COLORS['primary_dark'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Legend
        row = 3
        legend = [
            ('R', 'Responsible - Does the work', COLORS['success']),
            ('A', 'Accountable - Final authority', COLORS['danger']),
            ('C', 'Consulted - Provides input', COLORS['info']),
            ('I', 'Informed - Kept in the loop', COLORS['neutral']),
        ]

        for code, desc, color in legend:
            code_cell = ws.cell(row=row, column=1)
            code_cell.value = code
            code_cell.font = Font(bold=True, color=COLORS['white'])
            code_cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            code_cell.alignment = Alignment(horizontal='center', vertical='center')

            desc_cell = ws.cell(row=row, column=2)
            desc_cell.value = desc
            desc_cell.alignment = Alignment(horizontal='left', vertical='center')

            row += 1

        # RACI Table
        row += 1
        activities = ['Project Kickoff', 'Requirements', 'Design Review', 'Development',
                     'Testing', 'UAT', 'Deployment', 'Hypercare']
        roles = ['PM', 'BA', 'Architect', 'Dev Lead', 'QA Lead', 'Client PM', 'DevOps']

        # Headers
        ws.cell(row=row, column=1).value = 'Activity'
        ws.cell(row=row, column=1).font = Font(bold=True, color=COLORS['white'])
        ws.cell(row=row, column=1).fill = PatternFill(start_color=COLORS['primary'],
                                                     end_color=COLORS['primary'], fill_type='solid')

        for col, role in enumerate(roles, 2):
            cell = ws.cell(row=row, column=col)
            cell.value = role
            cell.font = Font(bold=True, color=COLORS['white'], size=10)
            cell.fill = PatternFill(start_color=COLORS['primary'],
                                  end_color=COLORS['primary'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions[get_column_letter(col)].width = 12

        row += 1

        # Sample assignments
        raci_data = {
            'Project Kickoff': ['A', 'C', 'C', 'C', 'C', 'C', 'I'],
            'Requirements': ['C', 'A', 'C', 'C', 'I', 'C', 'I'],
            'Design Review': ['C', 'C', 'A', 'C', 'I', 'I', 'I'],
            'Development': ['C', 'I', 'C', 'A', 'I', 'I', 'I'],
            'Testing': ['C', 'I', 'I', 'C', 'A', 'I', 'I'],
            'UAT': ['C', 'I', 'I', 'I', 'C', 'A', 'I'],
            'Deployment': ['C', 'I', 'I', 'I', 'I', 'C', 'A'],
            'Hypercare': ['A', 'I', 'I', 'C', 'I', 'C', 'C'],
        }

        for activity in activities:
            ws.cell(row=row, column=1).value = activity
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=1).fill = PatternFill(start_color=COLORS['header_light'],
                                                         end_color=COLORS['header_light'],
                                                         fill_type='solid')

            for col, role in enumerate(roles, 2):
                assignment = raci_data.get(activity, ['I'] * len(roles))[col - 2]
                cell = ws.cell(row=row, column=col)
                cell.value = assignment
                cell.font = Font(bold=True, size=11)

                color_map = {'R': COLORS['success'], 'A': COLORS['danger'],
                            'C': COLORS['info'], 'I': COLORS['neutral']}
                cell.fill = PatternFill(start_color=color_map.get(assignment, COLORS['white']),
                                       end_color=color_map.get(assignment, COLORS['white']),
                                       fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')

            row += 1

        ExcelFormatter.freeze_panes(ws, 9)

    @staticmethod
    def create_weekly_status(workbook, project_info: Dict):
        """Create PMO-style weekly status tracking sheet."""
        ws = workbook.create_sheet('11_Weekly_Status', 11)

        # Title
        ws.merge_cells('A1:H1')
        cell = ws['A1']
        cell.value = 'WEEKLY STATUS TRACKER'
        cell.font = Font(bold=True, size=14, color=COLORS['white'])
        cell.fill = PatternFill(start_color=COLORS['primary_dark'],
                              end_color=COLORS['primary_dark'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Headers
        headers = ['Week', 'Start Date', 'End Date', 'Planned %', 'Actual %', 'Variance %', 'Status', 'Comments']
        row = 3
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True, size=11, color=COLORS['white'])
            cell.fill = PatternFill(start_color=COLORS['primary'],
                                  end_color=COLORS['primary'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Weekly data
        row = 4
        duration = project_info.get('duration_weeks', 16)
        start_date = datetime.strptime(str(project_info.get('start_date', '2025-01-01')), '%Y-%m-%d')

        for week in range(1, min(duration + 1, 25)):
            week_start = start_date + timedelta(weeks=week - 1)
            week_end = week_start + timedelta(days=6)

            PMOWorkbookOptimizer._format_value(ws, row, 1, f'Week {week}')
            PMOWorkbookOptimizer._format_value(ws, row, 2, week_start.strftime('%Y-%m-%d'))
            PMOWorkbookOptimizer._format_value(ws, row, 3, week_end.strftime('%Y-%m-%d'))

            planned = int((week / duration) * 100)
            ws.cell(row=row, column=4).value = f'{planned}%'
            ws.cell(row=row, column=4).alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(row=row, column=5).value = '0%'
            ws.cell(row=row, column=5).alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(row=row, column=6).value = '0%'
            ws.cell(row=row, column=6).alignment = Alignment(horizontal='center', vertical='center')

            status_cell = ws.cell(row=row, column=7)
            status_cell.value = 'Pending'
            status_cell.fill = PatternFill(start_color=COLORS['neutral'],
                                         end_color=COLORS['neutral'], fill_type='solid')
            status_cell.alignment = Alignment(horizontal='center', vertical='center')

            PMOWorkbookOptimizer._format_value(ws, row, 8, '')
            row += 1

        col_widths = {'A': 12, 'B': 14, 'C': 14, 'D': 12, 'E': 12, 'F': 12, 'G': 12, 'H': 20}
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        ExcelFormatter.freeze_panes(ws, 4)
        ExcelFormatter.add_filter(ws, 'A3', 'H100')
