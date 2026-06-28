"""
Workbook Generator Service for Project Aura.
Generates complete Excel workbooks with all project planning sheets.
"""

import logging
import os
from datetime import datetime
from typing import Dict, Any
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.styles import PatternFill, Alignment

from services.excel_formatter import ExcelFormatter
from services.project_plan_engine import ProjectPlanEngine

logger = logging.getLogger(__name__)


class WorkbookGenerator:
    """
    Service for generating complete Excel workbooks.
    """

    def __init__(self, project_info: Dict, db_summary: Dict):
        """
        Initialize workbook generator.
        
        Args:
            project_info: Project information
            db_summary: Database project summary
        """
        self.project_info = project_info
        self.db_summary = db_summary
        self.workbook = None
        self.project_type = project_info.get('project_type', 'Unknown')
        self.client_name = project_info.get('client_name', 'Client')

    def generate(self, output_path: str) -> bool:
        """
        Generate complete workbook.
        
        Args:
            output_path: Path to save workbook
        
        Returns:
            True if successful
        """
        try:
            self.workbook = Workbook()
            self.workbook.remove(self.workbook.active)  # Remove default sheet
            
            # Generate all sheets in order
            self._create_project_details()
            self._create_project_charter()
            self._create_assumptions()
            self._create_staffing_plan()
            self._create_project_plan()
            self._create_wbs()
            self._create_milestones()
            self._create_dependencies()
            self._create_risk_register()
            self._create_raci_matrix()
            self._create_leave_planner()
            self._create_project_tracker()
            self._create_holiday_calendar()
            self._create_dashboard()
            
            # Save workbook
            self.workbook.save(output_path)
            logger.info(f"Workbook generated successfully: {output_path}")
            return True
        
        except Exception as e:
            logger.error(f"Error generating workbook: {str(e)}")
            return False

    def _create_sheet(self, title: str, color: str = '366092') -> Any:
        """Create a new sheet with formatting."""
        ws = self.workbook.create_sheet(title)
        ExcelFormatter.set_sheet_color(ws, color)
        return ws

    def _create_project_details(self):
        """Create Project Details sheet."""
        ws = self._create_sheet('01_Project_Details')
        
        # Title
        ExcelFormatter.format_title(ws, 1, 1, f"{self.client_name} - Project Details")
        ExcelFormatter.merge_cells(ws, 'A1', 'D1')
        
        # Project Information
        row = 3
        ExcelFormatter.format_label(ws, row, 1, 'Project Name')
        ExcelFormatter.format_value(ws, row, 2, self.project_info.get('project_name', 'Project'))
        
        row += 1
        ExcelFormatter.format_label(ws, row, 1, 'Project Type')
        ExcelFormatter.format_value(ws, row, 2, self.project_type)
        
        row += 1
        ExcelFormatter.format_label(ws, row, 1, 'Client Name')
        ExcelFormatter.format_value(ws, row, 2, self.client_name)
        
        row += 1
        ExcelFormatter.format_label(ws, row, 1, 'Start Date')
        ExcelFormatter.format_value(ws, row, 2, self.project_info.get('start_date'))
        
        row += 1
        ExcelFormatter.format_label(ws, row, 1, 'Duration (Weeks)')
        ExcelFormatter.format_value(ws, row, 2, self.project_info.get('duration_weeks'))
        
        row += 1
        ExcelFormatter.format_label(ws, row, 1, 'Team Size')
        ExcelFormatter.format_value(ws, row, 2, self.project_info.get('team_size'))
        
        row += 1
        ExcelFormatter.format_label(ws, row, 1, 'Delivery Model')
        ExcelFormatter.format_value(ws, row, 2, self.project_info.get('delivery_model'))
        
        row += 2
        ExcelFormatter.format_label(ws, row, 1, 'Project Scope')
        ws.cell(row=row, column=2).value = self.project_info.get('scope', 'N/A')
        
        # Set column widths
        ExcelFormatter.set_column_widths(ws, {'A': 20, 'B': 40, 'C': 20, 'D': 20})
        ExcelFormatter.freeze_panes(ws, 2)

    def _create_project_charter(self):
        """Create Project Charter sheet."""
        ws = self._create_sheet('02_Project_Charter', 'C5504F')
        
        ExcelFormatter.format_title(ws, 1, 1, 'Project Charter')
        ExcelFormatter.merge_cells(ws, 'A1', 'D1')
        
        row = 3
        headers = ['Element', 'Description', 'Owner', 'Approval']
        for col, header in enumerate(headers, 1):
            ExcelFormatter.format_header_row(ws, row, [col])
            ws.cell(row=row, column=col).value = header
        
        row += 1
        charter_elements = [
            ('Project Purpose', self.project_info.get('scope', 'N/A'), 'PM', 'Pending'),
            ('Success Criteria', 'Deliver on schedule and within budget', 'PM', 'Pending'),
            ('Key Stakeholders', self.client_name, 'PM', 'Pending'),
            ('High-Level Risks', 'Resource availability, scope creep', 'PM', 'Pending'),
            ('Assumptions', 'Team available full-time, stable requirements', 'PM', 'Pending'),
        ]
        
        for element, description, owner, approval in charter_elements:
            ExcelFormatter.format_data_row(ws, row, [1, 2, 3, 4])
            ws.cell(row=row, column=1).value = element
            ws.cell(row=row, column=2).value = description
            ws.cell(row=row, column=3).value = owner
            ws.cell(row=row, column=4).value = approval
            row += 1
        
        ExcelFormatter.set_column_widths(ws, {'A': 20, 'B': 40, 'C': 15, 'D': 15})
        ExcelFormatter.freeze_panes(ws, 2)

    def _create_assumptions(self):
        """Create Assumptions sheet."""
        ws = self._create_sheet('03_Assumptions', '70AD47')
        
        ExcelFormatter.format_title(ws, 1, 1, 'Project Assumptions')
        ExcelFormatter.merge_cells(ws, 'A1', 'D1')
        
        row = 3
        headers = ['Assumption', 'Rationale', 'Impact if Invalid', 'Status']
        for col, header in enumerate(headers, 1):
            ExcelFormatter.format_header_row(ws, row, [col])
            ws.cell(row=row, column=col).value = header
        
        row += 1
        assumptions = [
            ('Team members dedicated full-time', 'Maximum productivity', 'Schedule delay', 'Active'),
            ('Requirements stable', 'Minimal rework', 'Scope creep', 'Active'),
            ('Stakeholder availability', 'Timely approvals', 'Decision delays', 'Active'),
            ('Technology stack approved', 'No framework changes', 'Architecture rework', 'Active'),
            ('Budget approved', 'No cost overruns', 'Project delay', 'Active'),
        ]
        
        for assumption, rationale, impact, status in assumptions:
            ExcelFormatter.format_data_row(ws, row, [1, 2, 3, 4])
            ws.cell(row=row, column=1).value = assumption
            ws.cell(row=row, column=2).value = rationale
            ws.cell(row=row, column=3).value = impact
            ws.cell(row=row, column=4).value = status
            row += 1
        
        ExcelFormatter.set_column_widths(ws, {'A': 25, 'B': 25, 'C': 25, 'D': 15})
        ExcelFormatter.freeze_panes(ws, 2)

    def _create_staffing_plan(self):
        """Create Staffing Plan sheet."""
        ws = self._create_sheet('04_Staffing_Plan', '4472C4')
        
        ExcelFormatter.format_title(ws, 1, 1, 'Project Staffing Plan')
        ExcelFormatter.merge_cells(ws, 'A1', 'F1')
        
        row = 3
        headers = ['Role', 'Count', 'Resource', 'Start Date', 'End Date', 'Allocation %']
        for col, header in enumerate(headers, 1):
            ExcelFormatter.format_header_row(ws, row, [col])
            ws.cell(row=row, column=col).value = header
        
        row += 1
        team_members = self.db_summary.get('team_members', [])
        start_date = self.project_info.get('start_date', 'TBD')
        duration_weeks = self.project_info.get('duration_weeks', 0)
        
        for member in team_members:
            ExcelFormatter.format_data_row(ws, row, [1, 2, 3, 4, 5, 6])
            ws.cell(row=row, column=1).value = member.get('role', '')
            ws.cell(row=row, column=2).value = member.get('count', 1)
            ws.cell(row=row, column=3).value = 'TBD'
            ws.cell(row=row, column=4).value = start_date
            ws.cell(row=row, column=5).value = start_date  # Will be calculated
            ws.cell(row=row, column=6).value = 1.0
            ExcelFormatter.format_percentage_column(ws, row, row, 6)
            row += 1
        
        ExcelFormatter.set_column_widths(ws, {'A': 20, 'B': 10, 'C': 20, 'D': 15, 'E': 15, 'F': 12})
        ExcelFormatter.freeze_panes(ws, 2)

    def _create_project_plan(self):
        """Create Project Plan sheet."""
        ws = self._create_sheet('05_Project_Plan', '8E7CC3')
        
        ExcelFormatter.format_title(ws, 1, 1, 'Detailed Project Plan')
        ExcelFormatter.merge_cells(ws, 'A1', 'F1')
        
        # Generate plan
        plan_data = ProjectPlanEngine.generate_project_plan(self.project_info)
        
        row = 3
        headers = ['Phase', 'Start Date', 'End Date', 'Duration (Days)', 'Status', 'Owner']
        for col, header in enumerate(headers, 1):
            ExcelFormatter.format_header_row(ws, row, [col])
            ws.cell(row=row, column=col).value = header
        
        row += 1
        for phase in plan_data.get('phases', []):
            ExcelFormatter.format_data_row(ws, row, [1, 2, 3, 4, 5, 6])
            ws.cell(row=row, column=1).value = phase.get('phase', '')
            ws.cell(row=row, column=2).value = phase.get('start_date', '')
            ws.cell(row=row, column=3).value = phase.get('end_date', '')
            ws.cell(row=row, column=4).value = phase.get('duration_days', 0)
            ws.cell(row=row, column=5).value = phase.get('status', 'Planned')
            ws.cell(row=row, column=6).value = 'PM'
            row += 1
        
        ExcelFormatter.set_column_widths(ws, {'A': 25, 'B': 15, 'C': 15, 'D': 15, 'E': 12, 'F': 12})
        ExcelFormatter.freeze_panes(ws, 2)

    def _create_wbs(self):
        """Create Work Breakdown Structure sheet."""
        ws = self._create_sheet('06_WBS', 'F4B183')
        
        ExcelFormatter.format_title(ws, 1, 1, 'Work Breakdown Structure (WBS)')
        ExcelFormatter.merge_cells(ws, 'A1', 'E1')
        
        row = 3
        headers = ['WBS Level', 'Work Package', 'Description', 'Duration (Days)', 'Effort (%)']
        for col, header in enumerate(headers, 1):
            ExcelFormatter.format_header_row(ws, row, [col])
            ws.cell(row=row, column=col).value = header
        
        row += 1
        plan_data = ProjectPlanEngine.generate_project_plan(self.project_info)
        
        for i, phase in enumerate(plan_data.get('phases', []), 1):
            ExcelFormatter.format_data_row(ws, row, [1, 2, 3, 4, 5])
            ws.cell(row=row, column=1).value = f"1.{i}"
            ws.cell(row=row, column=2).value = phase.get('phase', '')
            ws.cell(row=row, column=3).value = f"Execute {phase.get('phase', '').lower()}"
            ws.cell(row=row, column=4).value = phase.get('duration_days', 0)
            ws.cell(row=row, column=5).value = phase.get('duration_days', 0) / (self.project_info.get('duration_weeks', 1) * 5)
            ExcelFormatter.format_percentage_column(ws, row, row, 5)
            row += 1
        
        ExcelFormatter.set_column_widths(ws, {'A': 15, 'B': 25, 'C': 30, 'D': 15, 'E': 12})
        ExcelFormatter.freeze_panes(ws, 2)

    def _create_milestones(self):
        """Create Milestones sheet."""
        ws = self._create_sheet('07_Milestones', 'FF9800')
        
        ExcelFormatter.format_title(ws, 1, 1, 'Project Milestones')
        ExcelFormatter.merge_cells(ws, 'A1', 'E1')
        
        row = 3
        headers = ['Milestone', 'Date', 'Phase', 'Priority', 'Status']
        for col, header in enumerate(headers, 1):
            ExcelFormatter.format_header_row(ws, row, [col])
            ws.cell(row=row, column=col).value = header
        
        row += 1
        plan_data = ProjectPlanEngine.generate_project_plan(self.project_info)
        
        for milestone in plan_data.get('milestones', []):
            ExcelFormatter.format_data_row(ws, row, [1, 2, 3, 4, 5])
            ws.cell(row=row, column=1).value = milestone.get('milestone', '')
            ws.cell(row=row, column=2).value = milestone.get('date', '')
            ws.cell(row=row, column=3).value = milestone.get('phase', '')
            priority = milestone.get('priority', 'Medium')
            ws.cell(row=row, column=4).value = priority
            ws.cell(row=row, column=5).value = milestone.get('status', 'Planned')
            
            # Color code priority
            if priority == 'Critical':
                color = ExcelFormatter.COLORS['danger']
            elif priority == 'High':
                color = ExcelFormatter.COLORS['warning']
            else:
                color = None
            
            if color:
                ws.cell(row=row, column=4).fill = PatternFill(start_color=color, 
                                                               end_color=color, 
                                                               fill_type='solid')
            row += 1
        
        ExcelFormatter.set_column_widths(ws, {'A': 30, 'B': 15, 'C': 20, 'D': 12, 'E': 12})
        ExcelFormatter.freeze_panes(ws, 2)

    def _create_dependencies(self):
        """Create Dependencies sheet."""
        ws = self._create_sheet('08_Dependencies', '2196F3')
        
        ExcelFormatter.format_title(ws, 1, 1, 'Task Dependencies')
        ExcelFormatter.merge_cells(ws, 'A1', 'E1')
        
        row = 3
        headers = ['Task', 'Depends On', 'Type', 'Lead/Lag', 'Critical']
        for col, header in enumerate(headers, 1):
            ExcelFormatter.format_header_row(ws, row, [col])
            ws.cell(row=row, column=col).value = header
        
        row += 1
        plan_data = ProjectPlanEngine.generate_project_plan(self.project_info)
        
        for dep in plan_data.get('dependencies', []):
            ExcelFormatter.format_data_row(ws, row, [1, 2, 3, 4, 5])
            ws.cell(row=row, column=1).value = dep.get('task', '')
            ws.cell(row=row, column=2).value = dep.get('depends_on', '')
            ws.cell(row=row, column=3).value = dep.get('dependency_type', '')
            ws.cell(row=row, column=4).value = dep.get('lead_lag', 0)
            ws.cell(row=row, column=5).value = 'Yes' if dep in plan_data.get('critical_path', []) else 'No'
            row += 1
        
        ExcelFormatter.set_column_widths(ws, {'A': 25, 'B': 25, 'C': 20, 'D': 12, 'E': 10})
        ExcelFormatter.freeze_panes(ws, 2)

    def _create_risk_register(self):
        """Create Risk Register sheet."""
        ws = self._create_sheet('09_Risk_Register', 'E74C3C')
        
        ExcelFormatter.format_title(ws, 1, 1, 'Risk Register')
        ExcelFormatter.merge_cells(ws, 'A1', 'G1')
        
        row = 3
        headers = ['Risk ID', 'Risk Description', 'Probability', 'Impact', 'Severity', 'Mitigation', 'Owner']
        for col, header in enumerate(headers, 1):
            ExcelFormatter.format_header_row(ws, row, [col])
            ws.cell(row=row, column=col).value = header
        
        row += 1
        risks = self.db_summary.get('risks', [])
        
        for i, risk in enumerate(risks, 1):
            ExcelFormatter.format_data_row(ws, row, [1, 2, 3, 4, 5, 6, 7])
            ws.cell(row=row, column=1).value = f"R-{i:03d}"
            ws.cell(row=row, column=2).value = risk.get('risk_description', '')
            ws.cell(row=row, column=3).value = 0.5  # Default probability
            ws.cell(row=row, column=4).value = 0.5  # Default impact
            ExcelFormatter.format_percentage_column(ws, row, row, 3)
            ExcelFormatter.format_percentage_column(ws, row, row, 4)
            
            severity = risk.get('severity', 'Medium')
            ws.cell(row=row, column=5).value = severity
            ExcelFormatter.format_severity(ws, row, 5, severity)
            
            ws.cell(row=row, column=6).value = risk.get('mitigation', '')
            ws.cell(row=row, column=7).value = 'PM'
            row += 1
        
        ExcelFormatter.set_column_widths(ws, {'A': 10, 'B': 30, 'C': 12, 'D': 10, 'E': 12, 'F': 25, 'G': 12})
        ExcelFormatter.freeze_panes(ws, 2)

    def _create_raci_matrix(self):
        """Create RACI Matrix sheet."""
        ws = self._create_sheet('10_RACI_Matrix', '16A085')
        
        ExcelFormatter.format_title(ws, 1, 1, 'RACI Matrix')
        ExcelFormatter.merge_cells(ws, 'A1', 'E1')
        
        row = 3
        roles = ['PM', 'Tech Lead', 'Developer', 'QA', 'DevOps']
        
        ws.cell(row=row, column=1).value = 'Activity/Deliverable'
        for col, role in enumerate(roles, 2):
            ExcelFormatter.format_header_row(ws, row, [col])
            ws.cell(row=row, column=col).value = role
        
        row += 1
        activities = ['Project Initiation', 'Requirements Gathering', 'Design Review', 'Development', 
                     'Testing', 'Deployment', 'Support']
        
        raci_data = {
            'Project Initiation': ['R', 'C', '', '', ''],
            'Requirements Gathering': ['R', 'A', 'C', 'C', ''],
            'Design Review': ['C', 'R', 'C', 'C', 'C'],
            'Development': ['C', 'A', 'R', 'C', 'C'],
            'Testing': ['C', 'C', 'I', 'R', 'C'],
            'Deployment': ['R', 'A', 'C', 'C', 'R'],
            'Support': ['C', 'C', 'I', 'C', 'R']
        }
        
        for activity in activities:
            ExcelFormatter.format_data_row(ws, row, list(range(1, 7)))
            ws.cell(row=row, column=1).value = activity
            
            for col, value in enumerate(raci_data.get(activity, [''] * 5), 2):
                ws.cell(row=row, column=col).value = value
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
            
            row += 1
        
        # Add legend
        row += 1
        legend = [('R', 'Responsible'), ('A', 'Accountable'), ('C', 'Consulted'), ('I', 'Informed')]
        for abbrev, meaning in legend:
            ExcelFormatter.format_label(ws, row, 1, f"{abbrev} = {meaning}")
            row += 1
        
        ExcelFormatter.set_column_widths(ws, {'A': 25, 'B': 12, 'C': 12, 'D': 12, 'E': 12, 'F': 12})
        ExcelFormatter.freeze_panes(ws, 2)

    def _create_leave_planner(self):
        """Create Leave Planner sheet."""
        ws = self._create_sheet('11_Leave_Planner', 'F39C12')
        
        ExcelFormatter.format_title(ws, 1, 1, 'Team Leave Planner')
        ExcelFormatter.merge_cells(ws, 'A1', 'F1')
        
        row = 3
        headers = ['Resource', 'Leave Type', 'Start Date', 'End Date', 'Days', 'Approval Status']
        for col, header in enumerate(headers, 1):
            ExcelFormatter.format_header_row(ws, row, [col])
            ws.cell(row=row, column=col).value = header
        
        row += 1
        # Sample leave entries
        ws.cell(row=row, column=1).value = 'Team Member 1'
        ws.cell(row=row, column=2).value = 'Vacation'
        ws.cell(row=row, column=3).value = ''
        ws.cell(row=row, column=4).value = ''
        ws.cell(row=row, column=5).value = 0
        ws.cell(row=row, column=6).value = 'Pending'
        
        ExcelFormatter.set_column_widths(ws, {'A': 20, 'B': 15, 'C': 15, 'D': 15, 'E': 10, 'F': 15})
        ExcelFormatter.freeze_panes(ws, 2)

    def _create_project_tracker(self):
        """Create Project Tracker sheet."""
        ws = self._create_sheet('12_Project_Tracker', '3498DB')
        
        ExcelFormatter.format_title(ws, 1, 1, 'Project Tracker')
        ExcelFormatter.merge_cells(ws, 'A1', 'G1')
        
        row = 3
        headers = ['Item', 'Planned', 'Actual', 'Variance', 'Status', 'Notes', 'Owner']
        for col, header in enumerate(headers, 1):
            ExcelFormatter.format_header_row(ws, row, [col])
            ws.cell(row=row, column=col).value = header
        
        row += 1
        tracker_items = [
            ('Timeline', 'On Track', 'On Track', '0%', 'Green', 'All phases on schedule', 'PM'),
            ('Budget', 'On Track', 'On Track', '0%', 'Green', 'No overruns', 'PM'),
            ('Quality', 'On Track', 'On Track', '0%', 'Green', 'Defects within threshold', 'QA'),
            ('Team', 'On Track', 'On Track', '0%', 'Green', 'Full staffing', 'PM'),
        ]
        
        for item, planned, actual, variance, status, notes, owner in tracker_items:
            ExcelFormatter.format_data_row(ws, row, [1, 2, 3, 4, 5, 6, 7])
            ws.cell(row=row, column=1).value = item
            ws.cell(row=row, column=2).value = planned
            ws.cell(row=row, column=3).value = actual
            ws.cell(row=row, column=4).value = variance
            ws.cell(row=row, column=5).value = status
            ws.cell(row=row, column=6).value = notes
            ws.cell(row=row, column=7).value = owner
            row += 1
        
        ExcelFormatter.set_column_widths(ws, {'A': 15, 'B': 15, 'C': 15, 'D': 12, 'E': 10, 'F': 25, 'G': 12})
        ExcelFormatter.freeze_panes(ws, 2)

    def _create_holiday_calendar(self):
        """Create Holiday Calendar sheet."""
        ws = self._create_sheet('13_Holiday_Calendar', '9B59B6')
        
        ExcelFormatter.format_title(ws, 1, 1, 'Holiday Calendar 2025')
        ExcelFormatter.merge_cells(ws, 'A1', 'C1')
        
        row = 3
        headers = ['Date', 'Holiday', 'Impact on Schedule']
        for col, header in enumerate(headers, 1):
            ExcelFormatter.format_header_row(ws, row, [col])
            ws.cell(row=row, column=col).value = header
        
        row += 1
        holidays = [
            ('2025-01-01', 'New Year Day', 'No work'),
            ('2025-03-17', "St. Patrick's Day", 'Optional'),
            ('2025-05-26', 'Memorial Day', 'No work'),
            ('2025-07-04', 'Independence Day', 'No work'),
            ('2025-09-01', 'Labor Day', 'No work'),
            ('2025-11-27', 'Thanksgiving', 'No work'),
            ('2025-12-25', 'Christmas', 'No work'),
        ]
        
        for date, holiday, impact in holidays:
            ExcelFormatter.format_data_row(ws, row, [1, 2, 3])
            ws.cell(row=row, column=1).value = date
            ws.cell(row=row, column=2).value = holiday
            ws.cell(row=row, column=3).value = impact
            row += 1
        
        ExcelFormatter.set_column_widths(ws, {'A': 15, 'B': 25, 'C': 25})
        ExcelFormatter.freeze_panes(ws, 2)

    def _create_dashboard(self):
        """Create Dashboard sheet."""
        ws = self._create_sheet('14_Dashboard', '27AE60')
        
        ExcelFormatter.format_title(ws, 1, 1, f'{self.client_name} - Project Dashboard')
        ExcelFormatter.merge_cells(ws, 'A1', 'H1')
        
        # Key metrics
        row = 3
        ExcelFormatter.format_label(ws, row, 1, 'Project Status Summary')
        
        row += 2
        metrics = [
            ('Project Type', self.project_type),
            ('Start Date', self.project_info.get('start_date', 'TBD')),
            ('Duration (Weeks)', self.project_info.get('duration_weeks', 0)),
            ('Team Size', self.project_info.get('team_size', 0)),
            ('Overall Status', 'On Track'),
            ('Schedule Health', '100%'),
        ]
        
        for label, value in metrics:
            ExcelFormatter.format_label(ws, row, 1, label)
            ExcelFormatter.format_value(ws, row, 2, value, bold=True)
            row += 1
        
        # Phase progress
        row += 2
        ExcelFormatter.format_label(ws, row, 1, 'Phase Progress')
        
        row += 1
        plan_data = ProjectPlanEngine.generate_project_plan(self.project_info)
        for phase in plan_data.get('phases', [])[:5]:  # Show first 5 phases
            ExcelFormatter.format_label(ws, row, 1, phase.get('phase', ''))
            ExcelFormatter.format_value(ws, row, 2, 0)  # Progress %
            ExcelFormatter.format_percentage_column(ws, row, row, 2)
            row += 1
        
        ExcelFormatter.set_column_widths(ws, {chr(65+i): 18 for i in range(8)})
        ExcelFormatter.freeze_panes(ws, 2)

    def _add_chart_to_sheet(self, ws, chart, position='D3'):
        """Add a chart to a worksheet."""
        ws.add_chart(chart, position)
